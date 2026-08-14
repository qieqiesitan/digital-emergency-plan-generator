"""风险分级管控清单：层级展平、默认管控层级、Excel 台账导出、公开脱敏。"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def default_control_level(mapping: dict, current_level: str | None) -> str:
    """按现有风险等级查默认管控层级；未知/缺失回退「岗位」。"""
    return mapping.get(current_level or "", "岗位")


def flatten_rows(zones: list, mapping: dict) -> list[dict]:
    """把 分区→风险点→单元→事件 树展平为清单行（含 zone_id/object_id 供筛选）。"""
    rows = []
    for z in zones:
        for obj in z.objects or []:
            for unit in obj.units or []:
                for ev in unit.events or []:
                    rows.append(_row(z, obj, unit, ev, mapping))
            for ev in obj.events or []:
                rows.append(_row(z, obj, None, ev, mapping))
    return rows


def _row(z, obj, unit, ev, mapping) -> dict:
    measures = "；".join(
        f"{m.measure_category}:{m.description}" for m in (ev.measures or [])) or "-"
    return {
        "zone_id": z.id, "object_id": obj.id,
        "zone": z.name, "object": obj.name, "unit": unit.name if unit else "-",
        "accident": ev.accident_type, "inherent": ev.inherent_risk_level or ev.risk_level or "-",
        "current": ev.risk_level or "-",
        "control_level": ev.control_level or default_control_level(mapping, ev.risk_level),
        "measures": measures, "unit_name": obj.responsible_unit or "-",
        "person": obj.responsible_person or "-", "phone": obj.contact_phone or "-",
    }


_COLUMN_MAP = {
    "分区": "zone", "风险点": "object", "单元": "unit", "事故类型": "accident",
    "固有等级": "inherent", "现有等级": "current", "管控层级": "control_level",
    "管控措施": "measures", "责任单位": "unit_name", "责任人": "person", "联系电话": "phone",
}

RISK_LEVEL_ORDER = ["低", "一般", "较大", "重大"]
CONTROL_LEVEL_ORDER = ["岗位", "班组", "部门", "企业"]


def _style_header_row(ws, row_idx: int) -> None:
    """给指定表头行加粗 + 浅灰底纹。"""
    for c in ws[row_idx]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="EEF2F7")


def build_ledger_workbook(rows: list[dict]) -> Workbook:
    """把英文键行（flatten_rows 输出）写入中文表头台账；sheet2 为等级/层级汇总。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "风险管控清单"
    headers = list(_COLUMN_MAP)
    ws.append(headers)
    _style_header_row(ws, 1)
    for r in rows:
        ws.append([r[_COLUMN_MAP[h]] for h in headers])

    ws2 = wb.create_sheet("等级层级汇总")
    ws2.append(["固有等级", "数量"])
    for level in RISK_LEVEL_ORDER:
        ws2.append([level, sum(1 for r in rows if r.get("inherent") == level)])
    ws2.append([None, None])  # 空行分隔两个汇总区
    ws2.append(["管控层级", "数量"])
    for level in CONTROL_LEVEL_ORDER:
        ws2.append([level, sum(1 for r in rows if r.get("control_level") == level)])
    _style_header_row(ws2, 1)
    _style_header_row(ws2, 7)
    return wb


PUBLIC_FIELDS = ["zone", "object", "unit", "accident", "inherent", "current",
                 "control_level", "measures", "unit_name"]


def desensitize(rows: list[dict]) -> list[dict]:
    """公开脱敏：仅保留 PUBLIC_FIELDS，不含 person/phone/内部键。"""
    return [{k: r.get(k) for k in PUBLIC_FIELDS} for r in rows]
