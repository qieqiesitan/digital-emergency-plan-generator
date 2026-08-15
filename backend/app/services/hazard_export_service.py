"""隐患台账与监管上报 Excel 导出服务（任务 11，§12/§14）。

openpyxl 纯函数：路由层负责查询 ORM 记录并解析名称/标签，本模块只做
「记录 → Workbook」的展示转换，便于直接单测（与 risk_control_list_service
的 `build_ledger_workbook` 先例同型）。

- `build_ledger_workbook`：企业内台账（含敏感字段），3 sheet——
  sheet1 台账、sheet2 超期清单（rectifying 且 deadline < 今天）、
  sheet3 重大隐患（level == major）。
- `build_report_workbook`：监管上报台账（脱敏白名单 8 列，不含责任人姓名/
  联系方式/照片）。
- `resolve_department_name`：经 enterprise_members.org_node_id 定位组织节点、
  向上找部门节点名（监管上报「责任单位」推导；无部门祖先缺省「—」）。
"""

from datetime import date, datetime
import json
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


# 台账 sheet1：按模型业务字段合理选取（docstring 见 build_ledger_workbook）。
LEDGER_HEADERS = [
    "编号", "标题", "描述", "隐患类型", "等级", "判定依据", "状态", "来源类型",
    "风险点", "管控措施", "位置", "照片", "治理方案", "整改期限", "整改责任人",
    "复查人", "登记人", "登记时间", "闭环时间",
]

# 台账 sheet2 超期清单：定位 + 超期天数。
OVERDUE_HEADERS = ["编号", "标题", "等级", "状态", "整改期限", "整改责任人", "超期天数"]

# 台账 sheet3 重大隐患专表。
MAJOR_HEADERS = ["编号", "标题", "等级", "状态", "整改期限", "判定依据", "整改责任人", "登记时间"]

# 监管上报台账（脱敏）：契约 §12 白名单字段。
REPORT_HEADERS = ["编号", "名称", "位置", "等级", "判定依据", "整改期限", "责任单位", "整改进度"]


def _style_header_row(ws, row_idx: int) -> None:
    """给指定表头行加粗 + 浅灰底纹（与 risk_control_list_service 同款）。"""
    for c in ws[row_idx]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="EEF2F7")


def _cell(value) -> str:
    """单元格文本兜底：None → '-'；日期/时间 → ISO；JSON 对象 → 中文 JSON 串。"""
    if value is None:
        return "-"
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return str(value)


def _name_or(mapping: dict, key, fallback) -> str:
    """名称解析：映射命中取名称；未命中回退原始 id；都为空 → '-'。"""
    if key:
        return mapping.get(key) or str(key)
    return "-"


def build_ledger_workbook(
    records,
    *,
    object_names: Optional[dict] = None,
    measure_names: Optional[dict] = None,
    user_names: Optional[dict] = None,
    status_labels: Optional[dict] = None,
    source_labels: Optional[dict] = None,
    hazard_type_labels: Optional[dict] = None,
    today: Optional[date] = None,
) -> Workbook:
    """企业内台账 Workbook（含敏感字段，仅限企业内使用），3 sheet。

    台账字段按模型列合理选取（说明）：
    - 业务字段全量：code/title/description/hazard_type/level/grading_basis/status/
      source_type/location/photo_urls/rectification_plan/deadline/created_at/closed_at；
    - 关联对象解析为名称：object_id → 风险点名、measure_id → 管控措施描述、
      rectification_user_id/reviewer_user_id/created_by → 用户姓名（映射未命中
      时保留原始 id，便于回查）；
    - 省略纯内部/审计列：id/enterprise_id/source_task_id/source_item_id/
      level_source/updated_at（避免台账噪音，不外泄排查任务内部关联）。
    等级列保留模型码值（major/general，与 API 一致）；状态/来源/隐患类型列
    优先取数据字典中文标签（record_status_label/source_type/hazard_type），
    未命中回退原始码值。

    sheet2 超期清单：rectifying 且 deadline < today（自然日口径，含超期天数）。
    sheet3 重大隐患：level == major 全量（不分状态）。
    """
    object_names = object_names or {}
    measure_names = measure_names or {}
    user_names = user_names or {}
    status_labels = status_labels or {}
    source_labels = source_labels or {}
    hazard_type_labels = hazard_type_labels or {}
    today = today or date.today()

    wb = Workbook()
    ws = wb.active
    ws.title = "台账"
    ws.append(LEDGER_HEADERS)
    _style_header_row(ws, 1)
    for r in records:
        plan = r.rectification_plan
        ws.append([
            r.code,
            r.title,
            r.description,
            hazard_type_labels.get(r.hazard_type) if r.hazard_type else "-",
            r.level or "-",
            r.grading_basis or "-",
            status_labels.get(r.status, r.status),
            source_labels.get(r.source_type, r.source_type),
            _name_or(object_names, r.object_id, r.object_id),
            _name_or(measure_names, r.measure_id, r.measure_id),
            r.location or "-",
            "、".join(r.photo_urls or []) or "-",
            _cell(plan) if plan else "-",
            _cell(r.deadline),
            _name_or(user_names, r.rectification_user_id, r.rectification_user_id),
            _name_or(user_names, r.reviewer_user_id, r.reviewer_user_id),
            _name_or(user_names, r.created_by, r.created_by),
            _cell(r.created_at),
            _cell(r.closed_at),
        ])

    ws2 = wb.create_sheet("超期清单")
    ws2.append(OVERDUE_HEADERS)
    _style_header_row(ws2, 1)
    overdue = [
        r for r in records
        if r.status == "rectifying" and r.deadline and r.deadline < today
    ]
    for r in overdue:
        ws2.append([
            r.code,
            r.title,
            r.level or "-",
            status_labels.get(r.status, r.status),
            _cell(r.deadline),
            _name_or(user_names, r.rectification_user_id, r.rectification_user_id),
            (today - r.deadline).days,
        ])

    ws3 = wb.create_sheet("重大隐患")
    ws3.append(MAJOR_HEADERS)
    _style_header_row(ws3, 1)
    for r in [r for r in records if r.level == "major"]:
        ws3.append([
            r.code,
            r.title,
            r.level or "-",
            status_labels.get(r.status, r.status),
            _cell(r.deadline),
            r.grading_basis or "-",
            _name_or(user_names, r.rectification_user_id, r.rectification_user_id),
            _cell(r.created_at),
        ])
    return wb


def resolve_department_name(node_id: Optional[str], node_map: dict) -> str:
    """成员组织节点 → 部门节点名（监管上报「责任单位」推导）。

    enterprise_members.org_node_id 指向组织树节点（部门/班组/岗位），沿
    parent_id 向上找最近 type == "dept" 的节点取名称；无组织归属、节点缺失或
    无部门祖先时返回「—」（监管上报缺省口径，不编造单位）。
    """
    if not node_id or node_id not in node_map:
        return "—"
    cur = node_id
    seen = set()
    while cur and cur in node_map and cur not in seen:
        seen.add(cur)
        node = node_map[cur]
        if node.get("type") == "dept":
            return node.get("name") or "—"
        cur = node.get("parent_id")
    return "—"


def build_report_workbook(
    records,
    *,
    org_dept_map: Optional[dict] = None,
    progress_map: Optional[dict] = None,
) -> Workbook:
    """监管上报台账 Workbook（脱敏）：8 列白名单，不含责任人姓名/联系方式/照片。

    字段（§12）：编号/名称/位置/等级/判定依据/整改期限/责任单位/整改进度。
    - 责任单位：路由层经 enterprise_members/org 节点推导后传入
      {rectification_user_id: 部门名}，本函数缺省「—」；
    - 整改进度：路由层传 {record_id: 最近整改 content 或状态标签}，缺省「—」；
    - 等级列保留模型码值（major/general，与 API 一致），监管侧按码值对照。
    排序：创建时间倒序（新登记在前）。
    """
    org_dept_map = org_dept_map or {}
    progress_map = progress_map or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "监管上报台账"
    ws.append(REPORT_HEADERS)
    _style_header_row(ws, 1)
    ordered = sorted(records, key=lambda r: r.created_at or datetime.min, reverse=True)
    for r in ordered:
        ws.append([
            r.code,
            r.title,
            r.location or "-",
            r.level or "-",
            r.grading_basis or "-",
            _cell(r.deadline),
            org_dept_map.get(r.rectification_user_id) or "—",
            progress_map.get(r.id) or "—",
        ])
    return wb
