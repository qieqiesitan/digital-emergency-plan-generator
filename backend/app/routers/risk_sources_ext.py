import json
import io
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from Crypto.Cipher import AES
from Crypto.Util.Padding import pad, unpad
import httpx

from app.database import get_db
from app.models.user import User
from app.models.enterprise import Enterprise, RiskSource, AIConfig
from app.schemas.risk_source import RiskSourceCreate, RiskSourceResponse
from app.schemas.common import ApiResponse
from app.dependencies import get_current_user
from app.config import settings

router = APIRouter(prefix="/enterprises", tags=["Risk Sources Extended"])

# --- Risk level matrix (from PRD-02) ---
RISK_MATRIX = {
    ("高", "高"): "重大",
    ("高", "中"): "重大",
    ("中", "高"): "重大",
    ("高", "低"): "较大",
    ("低", "高"): "较大",
    ("中", "中"): "较大",
    ("中", "低"): "一般",
    ("低", "中"): "一般",
    ("低", "低"): "低",
}

def _calc_risk_level(likelihood: str, severity: str) -> str:
    return RISK_MATRIX.get((likelihood or "中", severity or "中"), "一般")

def _cats_to_str(categories: list[str] | None) -> str:
    if not categories:
        return ""
    return ",".join(c for c in categories if c)

# --- AI helpers ---
def _decrypt_api_key(hex_str: str) -> str:
    key = settings.ENCRYPTION_KEY.encode()[:32].ljust(32, b"\0")
    cipher = AES.new(key, AES.MODE_ECB)
    return unpad(cipher.decrypt(bytes.fromhex(hex_str)), 16).decode()

PRESET_RISK_CATEGORIES = [
    "火灾", "爆炸", "触电", "中毒窒息", "机械伤害",
    "高处坠落", "物体打击", "车辆伤害", "淹溺", "坍塌",
    "锅炉爆炸", "容器爆炸",
]

async def _call_llm_nonstream(messages: list[dict], ai_config: AIConfig) -> str:
    try:
        api_key = _decrypt_api_key(ai_config.api_key_encrypted)
    except Exception:
        raise HTTPException(500, "AI 配置密钥解密失败")
    base = ai_config.base_url or {
        "openai": "https://api.openai.com/v1",
        "qwen": "https://dashscope.aliyuncs.com/compatible-mode/v1",
        "deepseek": "https://api.deepseek.com/v1",
    }.get(ai_config.provider, "")
    payload = {
        "model": ai_config.model_name,
        "messages": messages,
        "temperature": ai_config.temperature,
        "max_tokens": ai_config.max_tokens,
        "top_p": ai_config.top_p,
        "stream": False,
    }
    async with httpx.AsyncClient(timeout=120) as client:
        resp = await client.post(
            f"{base}/chat/completions",
            json=payload,
            headers={"Authorization": f"Bearer {api_key}"},
        )
        if resp.status_code != 200:
            if resp.status_code == 401:
                raise HTTPException(500, "AI API Key 无效或已过期，请在系统设置中重新配置 AI 模型")
            raise HTTPException(500, f"AI 调用失败: HTTP {resp.status_code}")
        data = resp.json()
        return data["choices"][0]["message"]["content"]

async def _get_enterprise_data(enterprise_id: str, user_id: str, db: AsyncSession) -> dict:
    result = await db.execute(
        select(Enterprise).where(Enterprise.id == enterprise_id, Enterprise.user_id == user_id)
    )
    ent = result.scalar_one_or_none()
    if not ent:
        raise HTTPException(404, "企业不存在")
    return {
        "name": ent.name,
        "industry": ent.industry or "",
        "business_scope": ent.business_scope or "",
        "building_overview": ent.building_overview or "",
        "employee_count": ent.employee_count,
        "address": ent.address or "",
    }

# --- Template download ---
@router.get("/{enterprise_id}/risk-sources/template")
async def download_risk_source_template(
    enterprise_id: str,
    current_user=Depends(get_current_user),
    db=Depends(get_db),
):
    await _get_enterprise_data(enterprise_id, current_user.id, db)

    wb = Workbook()
    ws = wb.active
    ws.title = "风险源模板"

    headers = ["风险类别", "风险名称", "位置", "风险描述", "可能性", "严重性", "控制措施"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 40

    # Data validation for risk categories
    cat_dv = DataValidation(type="list", formula1='"' + ",".join(PRESET_RISK_CATEGORIES) + '"', allow_blank=True)
    cat_dv.error = "请从列表中选择风险类别"
    cat_dv.errorTitle = "无效类别"
    ws.add_data_validation(cat_dv)
    cat_dv.add("A2:A1000")

    # Data validation for likelihood
    like_dv = DataValidation(type="list", formula1='"高,中,低"', allow_blank=True)
    like_dv.error = "请选择高、中或低"
    ws.add_data_validation(like_dv)
    like_dv.add("E2:E1000")

    # Data validation for severity
    sev_dv = DataValidation(type="list", formula1='"高,中,低"', allow_blank=True)
    sev_dv.error = "请选择高、中或低"
    ws.add_data_validation(sev_dv)
    sev_dv.add("F2:F1000")

    # Add a sample row
    sample = ["火灾", "原料仓库", "仓库东区", "大量可燃物堆积，电气线路老化风险", "中", "高", "定期巡检，安装烟雾报警和自动喷淋"]
    for col, val in enumerate(sample, 1):
        ws.cell(row=2, column=col, value=val)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=risk_sources_template.xlsx"},
    )

# --- Import from Excel ---
class ImportPreviewItem(BaseModel):
    row: int
    data: RiskSourceCreate
    errors: list[str] = []

class ImportPreviewResponse(BaseModel):
    items: list[ImportPreviewItem]
    valid_count: int
    error_count: int

@router.post("/{enterprise_id}/risk-sources/import", response_model=ApiResponse[ImportPreviewResponse])
async def import_risk_sources(
    enterprise_id: str,
    file: UploadFile = File(...),
    current_user=Depends(get_current_user),
    db=Depends(get_db),
):
    await _get_enterprise_data(enterprise_id, current_user.id, db)

    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "请上传 .xlsx 文件")

    contents = await file.read()
    wb = Workbook(io.BytesIO(contents))
    ws = wb.active
    items: list[ImportPreviewItem] = []

    valid_likelihood = {"高", "中", "低"}
    valid_severity = {"高", "中", "低"}
    valid_categories = set(PRESET_RISK_CATEGORIES)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(cell for cell in row):
            continue

        cats_raw = str(row[0] or "").strip()
        name = str(row[1] or "").strip()
        location = str(row[2] or "").strip() or None
        description = str(row[3] or "").strip() or None
        likelihood = str(row[4] or "").strip() or None
        severity = str(row[5] or "").strip() or None
        control_measures = str(row[6] or "").strip() or None

        errors: list[str] = []

        # Parse categories
        categories = [c.strip() for c in cats_raw.replace("，", ",").split(",") if c.strip()]
        if not categories:
            errors.append("风险类别不能为空")
        for c in categories:
            if c not in valid_categories:
                errors.append(f"无效的风险类别: {c}")

        if not name:
            errors.append("风险名称不能为空")

        if likelihood and likelihood not in valid_likelihood:
            errors.append(f"可能性无效: {likelihood}（应为高/中/低）")
        if severity and severity not in valid_severity:
            errors.append(f"严重性无效: {severity}（应为高/中/低）")

        item = ImportPreviewItem(
            row=row_idx,
            data=RiskSourceCreate(
                categories=categories if categories else [],
                name=name,
                location=location,
                description=description,
                likelihood=likelihood if likelihood in valid_likelihood else None,
                severity=severity if severity in valid_severity else None,
                control_measures=control_measures,
            ),
            errors=errors,
        )
        items.append(item)

    valid_count = sum(1 for i in items if not i.errors)
    error_count = sum(1 for i in items if i.errors)

    return ApiResponse(data=ImportPreviewResponse(items=items, valid_count=valid_count, error_count=error_count))

# --- AI question generation ---
class AIQuestionItem(BaseModel):
    id: str
    question: str

class AIQuestionsResponse(BaseModel):
    questions: list[AIQuestionItem]

@router.post("/{enterprise_id}/risk-sources/ai/questions", response_model=ApiResponse[AIQuestionsResponse])
async def get_risk_ai_questions(
    enterprise_id: str,
    current_user=Depends(get_current_user),
    db=Depends(get_db),
):
    ent_data = await _get_enterprise_data(enterprise_id, current_user.id, db)
    ai_config = (await db.execute(
        select(AIConfig).where(AIConfig.user_id == current_user.id)
    )).scalar_one_or_none()
    if not ai_config:
        raise HTTPException(400, "请先在系统设置中配置 AI 模型")

    # 查询该企业已有的风险源，用于去重
    existing_risks = (await db.execute(
        select(RiskSource).where(RiskSource.enterprise_id == enterprise_id)
    )).scalars().all()
    existing_risk_names = [r.name for r in existing_risks]
    existing_risk_categories = set()
    for r in existing_risks:
        if r.categories:
            for c in r.categories.split(","):
                existing_risk_categories.add(c.strip())

    existing_summary = ""
    if existing_risk_names:
        existing_summary = f"\n该企业已录入的风险源（请避免重复提问）：\n"
        for r in existing_risks:
            cats = r.categories or ""
            existing_summary += f"- {r.name}（类别：{cats}，位置：{r.location or "未指定"}）\n"

    system_prompt = (
        "你是一位持有国家注册安全工程师资格的专业应急预案编制专家，熟悉 GB/T 29639-2020 标准。"
        "你的任务是提出针对性问题以帮助识别企业尚未录入的风险源，"
        "必须严格避免对已录入风险源重复提问。"
    )
    user_prompt = f"""请根据以下企业信息，提出 3~5 个针对性问题以辅助识别该企业特有的、尚未录入的风险源。
问题应结合该企业的行业特点、生产工艺和建筑概况，使用简体中文。
**重要：已录入的风险源不要重复提问，问题应聚焦于尚未覆盖的风险领域。**

企业信息：
- 名称：{ent_data["name"]}
- 行业：{ent_data["industry"]}
- 经营范围：{ent_data["business_scope"]}
- 建筑/厂区概况：{ent_data["building_overview"]}
- 员工人数：{ent_data["employee_count"] or "未知"}
{existing_summary}
参考风险类别（全部）：{", ".join(PRESET_RISK_CATEGORIES)}

请以 JSON 格式输出，格式严格为：{{"questions": [{{"id": "q1", "question": "问题文本"}}]}}
只输出 JSON，不要任何解释或额外文本。"""

    try:
        raw = await _call_llm_nonstream(
            [{"role": "system", "content": system_prompt}, {"role": "user", "content": user_prompt}],
            ai_config,
        )
        # Strip markdown fences if present
        raw = raw.strip()
        if raw.startswith("```"):
            lines = raw.split("\n")
            raw = "\n".join(lines[1:]) if lines[0].startswith("```") else raw
            if raw.endswith("```"):
                raw = raw[:-3].strip()
        data = json.loads(raw)
        questions = [AIQuestionItem(**q) for q in data.get("questions", [])]
        return ApiResponse(data=AIQuestionsResponse(questions=questions))
    except HTTPException:
        raise
    except json.JSONDecodeError:
        raise HTTPException(500, f"AI 返回格式异常，无法解析 JSON: {raw[:200]}")
    except Exception as e:
        raise HTTPException(500, f"AI 调用失败: {str(e)}")

# --- AI generate risk sources ---
class AIAnswerInput(BaseModel):
    question_id: str
    question: str
    answer: str

class AIGenerateRequest(BaseModel):
    answers: list[AIAnswerInput]

class AIGenerateResponse(BaseModel):
    items: list[RiskSourceCreate]

@router.post("/{enterprise_id}/risk-sources/ai/generate", response_model=ApiResponse[AIGenerateResponse])
async def generate_risk_sources_ai(
    enterprise_id: str,
    body: AIGenerateRequest,
    current_user=Depends(get_current_user),
    db=Depends(get_db),
):
    ent_data = await _get_enterprise_data(enterprise_id, current_user.id, db)
    ai_config = (await db.execute(
        select(AIConfig).where(AIConfig.user_id == current_user.id)
    )).scalar_one_or_none()
    if not ai_config:
        raise HTTPException(400, "请先在系统设置中配置 AI 模型")

    # 查询已有风险源，在生成时也做去重参考
    existing_risks = (await db.execute(
        select(RiskSource).where(RiskSource.enterprise_id == enterprise_id)
    )).scalars().all()
    existing_risk_names = [r.name for r in existing_risks]

    existing_summary = ""
    if existing_risk_names:
        existing_summary = f"\n该企业已录入的风险源（严禁重复生成）：\n" + "\n".join(f"- {name}" for name in existing_risk_names)

    qa_text = "\n".join(f"Q: {a.question}\nA: {a.answer}" for a in body.answers)

    system_prompt = (
        "你是一位持有国家注册安全工程师资格的专业应急预案编制专家，熟悉 GB/T 29639-2020 标准。"
        "严禁生成与已录入风险源名称相同或实质重复的风险源。"
    )
    user_prompt = f"""请根据以下企业信息和用户回答，识别并列出该企业可能的风险源。

企业信息：
- 名称：{ent_data["name"]}
- 行业：{ent_data["industry"]}
- 经营范围：{ent_data["business_scope"]}
- 建筑/厂区概况：{ent_data["building_overview"]}
- 员工人数：{ent_data["employee_count"] or "未知"}
{existing_summary}
用户对风险调查的回答：
{qa_text}

请列出该企业尚未录入的主要风险源。每个风险源包含以下字段（请用简体中文）：
- categories: 风险类别列表，从以下选择：{", ".join(PRESET_RISK_CATEGORIES)}（可选多个，但优先选择最匹配的1-2个）
- name: 风险源名称（简明扼要，必须与已录入风险源名称不重复）
- location: 可能发生位置（根据企业信息推测）
- description: 风险描述（具体描述该风险的情景和后果）
- likelihood: 可能性（高/中/低）
- severity: 严重性（高/中/低）
- control_measures: 控制措施建议

请以 JSON 格式输出：{{"items": [{{"categories": ["火灾"], "name": "...", "location": "...", "description": "...", "likelihood": "中", "severity": "高", "control_measures": "..."}}]}}
只输出 JSON，不要任何解释。"""

    try:
        raw = await _call_llm_nonstream(
            [{"role": "system", "content": system_prompt}, {"role": "user", "content": user_prompt}],
            ai_config,
        )
        raw = raw.strip()
        if raw.startswith("```"):
            lines = raw.split("\n")
            raw = "\n".join(lines[1:]) if lines[0].startswith("```") else raw
            if raw.endswith("```"):
                raw = raw[:-3].strip()
        data = json.loads(raw)
        items = [RiskSourceCreate(**item) for item in data.get("items", [])]
        return ApiResponse(data=AIGenerateResponse(items=items))
    except HTTPException:
        raise
    except json.JSONDecodeError:
        raise HTTPException(500, f"AI 返回格式异常，无法解析 JSON: {raw[:200]}")
    except Exception as e:
        raise HTTPException(500, f"AI 调用失败: {str(e)}")

# --- Batch create ---
class BatchCreateRequest(BaseModel):
    items: list[RiskSourceCreate]

@router.post("/{enterprise_id}/risk-sources/batch", response_model=ApiResponse[list[RiskSourceResponse]], status_code=201)
async def batch_create_risk_sources(
    enterprise_id: str,
    body: BatchCreateRequest,
    current_user=Depends(get_current_user),
    db=Depends(get_db),
):
    await _get_enterprise_data(enterprise_id, current_user.id, db)

    if not body.items:
        raise HTTPException(400, "至少需要一个风险源")

    created: list[RiskSource] = []
    for item in body.items:
        likelihood = item.likelihood or "中"
        severity = item.severity or "中"
        risk_level = _calc_risk_level(likelihood, severity)
        r = RiskSource(
            enterprise_id=enterprise_id,
            categories=_cats_to_str(item.categories),
            name=item.name,
            location=item.location,
            description=item.description,
            likelihood=likelihood,
            severity=severity,
            risk_level=risk_level,
            control_measures=item.control_measures,
        )
        db.add(r)
        created.append(r)

    await db.commit()
    for r in created:
        await db.refresh(r)

    return ApiResponse(data=[RiskSourceResponse.model_validate(r) for r in created])
