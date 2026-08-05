import json

import io

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File

from fastapi.responses import StreamingResponse

from sqlalchemy.ext.asyncio import AsyncSession

from sqlalchemy import select

from pydantic import BaseModel

from openpyxl import Workbook

from openpyxl.styles import Font, PatternFill, Alignment

from openpyxl.worksheet.datavalidation import DataValidation


from app.database import get_db

from app.models.enterprise import Enterprise, EmergencyResource, AIConfig

from app.services.llm_client import llm_text_completion

from app.schemas.emergency_resource import EmergencyResourceCreate, EmergencyResourceResponse

from app.schemas.common import ApiResponse

from app.dependencies import get_current_user



router = APIRouter(prefix="/enterprises", tags=["Resources Extended"])



PRESET_INTERNAL_CATEGORIES = [

    "消防设施", "急救物资", "防护装备", "通讯设备",

    "照明设备", "破拆工具", "侦检设备", "堵漏器材",

]

PRESET_EXTERNAL_CATEGORIES = [

    "消防队", "医院", "公安机关", "安监部门", "环保部门",

]

ALL_RESOURCE_CATEGORIES = PRESET_INTERNAL_CATEGORIES + PRESET_EXTERNAL_CATEGORIES



async def _get_enterprise_data(enterprise_id: str, user_id: str, db: AsyncSession) -> dict:

    result = await db.execute(

        select(Enterprise).where(Enterprise.id == enterprise_id, Enterprise.user_id == user_id)

    )

    ent = result.scalar_one_or_none()

    if not ent:

        raise HTTPException(404, "企业不存在")

    return {

        "name": ent.name,

        "industry": ent.industry or "",

        "business_scope": ent.business_scope or "",

        "building_overview": ent.building_overview or "",

        "employee_count": ent.employee_count,

        "address": ent.address or "",
        "surrounding_info": ent.surrounding_info,

    }



# --- Template download ---

@router.get("/{enterprise_id}/resources/template")

async def download_resource_template(

    enterprise_id: str,

    current_user=Depends(get_current_user),

    db=Depends(get_db),

):

    await _get_enterprise_data(enterprise_id, current_user.id, db)



    wb = Workbook()

    ws = wb.active

    ws.title = "应急资源模板"



    headers = ["类别", "名称", "规格型号", "数量", "单位", "存放位置", "责任人", "联系电话", "是否外部", "外部地址", "距离(公里)"]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    header_font = Font(color="FFFFFF", bold=True, size=11)

    header_align = Alignment(horizontal="center", vertical="center")



    for col, h in enumerate(headers, 1):

        cell = ws.cell(row=1, column=col, value=h)

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = header_align



    widths = [16, 24, 20, 8, 8, 24, 12, 16, 10, 30, 12]

    for i, w in enumerate(widths, 1):

        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w



    # Data validation for category dropdown

    cat_formula = '""' + ",".join(ALL_RESOURCE_CATEGORIES) + '""'

    cat_dv = DataValidation(type="list", formula1=cat_formula, allow_blank=True)

    cat_dv.error = "请从列表中选择类别"

    cat_dv.errorTitle = "无效类别"

    ws.add_data_validation(cat_dv)

    cat_dv.add("A2:A1000")



    # Data validation for is_external

    ext_dv = DataValidation(type="list", formula1='"是,否"', allow_blank=True)

    ext_dv.error = "请选择是或否"

    ws.add_data_validation(ext_dv)

    ext_dv.add("I2:I1000")



    # Sample row

    sample = ["消防设施", "干粉灭火器", "MFZ/ABC8", 20, "个", "办公楼一楼走廊", "张三", "13800001111", "否", "", ""]

    for col, val in enumerate(sample, 1):

        ws.cell(row=2, column=col, value=val)



    output = io.BytesIO()

    wb.save(output)

    output.seek(0)



    return StreamingResponse(

        output,

        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

        headers={"Content-Disposition": "attachment; filename=emergency_resources_template.xlsx"},

    )



# --- Import from Excel ---

class ResourceImportPreviewItem(BaseModel):

    row: int

    data: EmergencyResourceCreate

    errors: list[str] = []



class ResourceImportPreviewResponse(BaseModel):

    items: list[ResourceImportPreviewItem]

    valid_count: int

    error_count: int



@router.post("/{enterprise_id}/resources/import", response_model=ApiResponse[ResourceImportPreviewResponse])

async def import_resources(

    enterprise_id: str,

    file: UploadFile = File(...),

    current_user=Depends(get_current_user),

    db=Depends(get_db),

):

    await _get_enterprise_data(enterprise_id, current_user.id, db)



    if not file.filename or not file.filename.endswith(".xlsx"):

        raise HTTPException(400, "请上传 .xlsx 文件")



    contents = await file.read()

    wb = Workbook(io.BytesIO(contents))

    ws = wb.active

    items: list[ResourceImportPreviewItem] = []

    valid_categories = set(ALL_RESOURCE_CATEGORIES)



    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

        if not row or not any(cell for cell in row):

            continue



        category = str(row[0] or "").strip()

        name = str(row[1] or "").strip()

        specification = str(row[2] or "").strip() or None

        qty_raw = row[3]

        try:

            quantity = int(qty_raw) if qty_raw is not None else 1

        except (ValueError, TypeError):

            quantity = 1

        unit = str(row[4] or "").strip() or None

        location = str(row[5] or "").strip() or None

        responsible_person = str(row[6] or "").strip() or None

        contact_phone = str(row[7] or "").strip() or None

        is_ext_raw = str(row[8] or "").strip()

        is_external = is_ext_raw == "是"

        external_address = str(row[9] or "").strip() or None if is_external else None

        dist_raw = row[10]

        try:

            external_distance_km = float(dist_raw) if dist_raw is not None and is_external else None

        except (ValueError, TypeError):

            external_distance_km = None



        errors: list[str] = []

        if not category:

            errors.append("类别不能为空")

        elif category not in valid_categories:

            errors.append(f"无效类别: {category}")

        if not name:

            errors.append("名称不能为空")

        if is_external and not external_address:

            errors.append("外部资源必须填写地址")



        item = ResourceImportPreviewItem(

            row=row_idx,

            data=EmergencyResourceCreate(

                category=category,

                name=name,

                specification=specification,

                quantity=quantity,

                unit=unit,

                location=location if not is_external else None,

                responsible_person=responsible_person,

                contact_phone=contact_phone,

                is_external=is_external,

                external_address=external_address,

                external_distance_km=external_distance_km,

            ),

            errors=errors,

        )

        items.append(item)



    valid_count = sum(1 for i in items if not i.errors)

    error_count = sum(1 for i in items if i.errors)



    return ApiResponse(data=ResourceImportPreviewResponse(items=items, valid_count=valid_count, error_count=error_count))



# --- AI question generation ---

class AIQuestionItem(BaseModel):

    id: str

    question: str



class AIQuestionsResponse(BaseModel):

    questions: list[AIQuestionItem]



@router.post("/{enterprise_id}/resources/ai/questions", response_model=ApiResponse[AIQuestionsResponse])

async def get_resource_ai_questions(

    enterprise_id: str,

    current_user=Depends(get_current_user),

    db=Depends(get_db),

):

    ent_data = await _get_enterprise_data(enterprise_id, current_user.id, db)

    ai_config = (await db.execute(

        select(AIConfig).where(AIConfig.user_id == current_user.id)

    )).scalar_one_or_none()

    if not ai_config:

        raise HTTPException(400, "请先在系统设置中配置 AI 模型")



    # 查询该企业已有的应急资源，用于去重

    existing_resources = (await db.execute(

        select(EmergencyResource).where(EmergencyResource.enterprise_id == enterprise_id)

    )).scalars().all()



    existing_summary = ""

    if existing_resources:

        existing_summary = "\n该企业已录入的应急资源（请避免重复提问）：\n"

        for r in existing_resources:

            is_ext = "外部" if r.is_external else "内部"

            existing_summary += f"- {r.name}（类别：{r.category}，{is_ext}，位置：{r.location or r.external_address or "未指定"}）\n"



    system_prompt = (

        "你是一位持有国家注册安全工程师资格的专业应急预案编制专家，熟悉 GB/T 29639-2020 标准。"

        "你的任务是提出针对性问题以帮助识别企业应急资源缺口，"

        "必须严格避免对已录入资源重复提问。"

    )

    user_prompt = f"""请根据以下企业信息，提出 3~5 个针对性问题以辅助识别该企业的应急资源缺口。

问题应结合该企业的行业特点、生产工艺和建筑概况，使用简体中文。

**重要：已录入的资源不要重复提问，问题应聚焦于尚未覆盖的资源缺口。**



企业信息：

- 名称：{ent_data["name"]}

- 行业：{ent_data["industry"]}

- 经营范围：{ent_data["business_scope"]}

- 建筑/厂区概况：{ent_data["building_overview"]}

- 员工人数：{ent_data["employee_count"] or "未知"}

{existing_summary}

内部资源类别参考：{", ".join(PRESET_INTERNAL_CATEGORIES)}

外部资源类别参考：{", ".join(PRESET_EXTERNAL_CATEGORIES)}



请以 JSON 格式输出：{{"questions": [{{"id": "q1", "question": "问题文本"}}]}}

只输出 JSON，不要任何解释。"""



    try:

        raw = await llm_text_completion(

            [{"role": "system", "content": system_prompt}, {"role": "user", "content": user_prompt}],

            ai_config,

        )

        raw = raw.strip()

        if raw.startswith("```"):

            lines = raw.split("\n")

            raw = "\n".join(lines[1:]) if lines[0].startswith("```") else raw

            if raw.endswith("```"):

                raw = raw[:-3].strip()

        data = json.loads(raw)

        questions = [AIQuestionItem(**q) for q in data.get("questions", [])]

        return ApiResponse(data=AIQuestionsResponse(questions=questions))

    except HTTPException:

        raise

    except json.JSONDecodeError:

        raise HTTPException(500, f"AI 返回格式异常，无法解析 JSON: {raw[:200]}")

    except Exception as e:

        raise HTTPException(500, f"AI 调用失败: {str(e)}")



# --- AI generate resources ---

class AIAnswerInput(BaseModel):

    question_id: str

    question: str

    answer: str



class AIGenerateRequest(BaseModel):

    answers: list[AIAnswerInput]



class AIGenerateResourceResponse(BaseModel):

    items: list[EmergencyResourceCreate]



@router.post("/{enterprise_id}/resources/ai/generate", response_model=ApiResponse[AIGenerateResourceResponse])

async def generate_resources_ai(

    enterprise_id: str,

    body: AIGenerateRequest,

    current_user=Depends(get_current_user),

    db=Depends(get_db),

):

    ent_data = await _get_enterprise_data(enterprise_id, current_user.id, db)

    ai_config = (await db.execute(

        select(AIConfig).where(AIConfig.user_id == current_user.id)

    )).scalar_one_or_none()

    if not ai_config:

        raise HTTPException(400, "请先在系统设置中配置 AI 模型")



    # 查询已有资源，生成时也做去重参考

    existing_resources = (await db.execute(

        select(EmergencyResource).where(EmergencyResource.enterprise_id == enterprise_id)

    )).scalars().all()

    existing_resource_names = [r.name for r in existing_resources]



    existing_summary = ""

    if existing_resource_names:

        existing_summary = "\n该企业已录入的应急资源（严禁重复生成）：\n" + "\n".join(f"- {name}" for name in existing_resource_names)



    qa_text = "\n".join(f"Q: {a.question}\nA: {a.answer}" for a in body.answers)
    # ponytail: format surrounding_info so AI can reference real nearby facilities
    surrounding_text = ""
    si = ent_data.get("surrounding_info")
    if si and isinstance(si, dict):
        parts = []
        nearby = si.get("nearby_units") or si.get("nearbyUnits") or []
        if nearby:
            lines_s = ["周边可用应急单位（外部资源请优先从下列真实单位中选取，严禁凭空编造名称和地址）："]
            for u in nearby:
                name = u.get("name", "")
                d = u.get("direction", "")
                dist = u.get("distance_m", 0) or 0
                risk = u.get("main_risk", "")
                lines_s.append(f"- {name}（{d}方向，约{dist}m） 应对风险：{risk}")
            parts.append("\n".join(lines_s))
        sensitive = si.get("sensitive_targets") or si.get("sensitiveTargets") or []
        if sensitive:
            lines_s = ["周边敏感目标（生成预案时需考虑的防护对象）："]
            for t in sensitive:
                name = t.get("name", "")
                d = t.get("direction", "")
                dist = t.get("distance_m", 0) or 0
                stype = t.get("type", "")
                lines_s.append(f"- {name}（{d}方向，约{dist}m） 类型：{stype}")
            parts.append("\n".join(lines_s))
        traffic = si.get("traffic_info", "")
        if traffic:
            parts.append(f"交通信息：{traffic}")
        if parts:
            surrounding_text = "\n\n" + "\n\n".join(parts) + "\n"





    system_prompt = (

        "你是一位持有国家注册安全工程师资格的专业应急预案编制专家，熟悉 GB/T 29639-2020 标准。"

        "严禁生成与已录入资源名称相同或实质重复的应急资源。"

    )

    user_prompt = f"""请根据以下企业信息和用户回答，列出该企业应配备的应急资源（包括内部物资和外部救援力量）。\n{surrounding_text}



企业信息：

- 名称：{ent_data["name"]}

- 行业：{ent_data["industry"]}

- 经营范围：{ent_data["business_scope"]}

- 建筑/厂区概况：{ent_data["building_overview"]}

- 员工人数：{ent_data["employee_count"] or "未知"}

{existing_summary}

用户对资源调查的回答：

{qa_text}



每个应急资源包含以下字段（请用简体中文）：

- category: 资源类别（内部从：{", ".join(PRESET_INTERNAL_CATEGORIES)} 选择；外部从：{", ".join(PRESET_EXTERNAL_CATEGORIES)} 选择）

- name: 资源名称（必须与已录入资源名称不重复）

- specification: 规格型号（内部资源填写）

- quantity: 建议数量（整数）

- unit: 单位（如：个、套、台、辆）

- location: 建议存放位置（内部资源填写）

- responsible_person: 责任人（可不填，留空字符串）

- contact_phone: 联系电话（可不填，留空字符串）

- is_external: 是否为外部资源（true/false）

- external_address: 外部地址（外部资源填写，内部资源留空）

- external_distance_km: 距离公里数（外部资源填写，内部资源留空）



请以 JSON 输出：{{"items": [{{"category": "消防设施", "name": "干粉灭火器", "specification": "MFZ/ABC8", "quantity": 20, "unit": "个", "location": "各车间门口", "responsible_person": "", "contact_phone": "", "is_external": false, "external_address": "", "external_distance_km": null}}]}}

只输出 JSON。"""



    try:

        raw = await llm_text_completion(

            [{"role": "system", "content": system_prompt}, {"role": "user", "content": user_prompt}],

            ai_config,

        )

        raw = raw.strip()

        if raw.startswith("```"):

            lines = raw.split("\n")

            raw = "\n".join(lines[1:]) if lines[0].startswith("```") else raw

            if raw.endswith("```"):

                raw = raw[:-3].strip()

        data = json.loads(raw)

        items = [EmergencyResourceCreate(**item) for item in data.get("items", [])]

        return ApiResponse(data=AIGenerateResourceResponse(items=items))

    except HTTPException:

        raise

    except json.JSONDecodeError:

        raise HTTPException(500, f"AI 返回格式异常，无法解析 JSON: {raw[:200]}")

    except Exception as e:

        raise HTTPException(500, f"AI 调用失败: {str(e)}")



# --- Batch create ---

class ResourceBatchCreateRequest(BaseModel):

    items: list[EmergencyResourceCreate]



@router.post("/{enterprise_id}/resources/batch", response_model=ApiResponse[list[EmergencyResourceResponse]], status_code=201)

async def batch_create_resources(

    enterprise_id: str,

    body: ResourceBatchCreateRequest,

    current_user=Depends(get_current_user),

    db=Depends(get_db),

):

    await _get_enterprise_data(enterprise_id, current_user.id, db)



    if not body.items:

        raise HTTPException(400, "至少需要一个资源")



    created: list[EmergencyResource] = []

    for item in body.items:

        r = EmergencyResource(

            enterprise_id=enterprise_id,

            **item.model_dump(exclude_none=True),

        )

        db.add(r)

        created.append(r)



    await db.commit()

    for r in created:

        await db.refresh(r)



    return ApiResponse(data=[EmergencyResourceResponse.model_validate(r) for r in created])
