import logging
from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.dependencies import get_current_user
from app.models.enterprise import Enterprise
from app.models.enterprise_org import EnterpriseMember
from app.models.user import User
from app.schemas.common import ApiResponse
from app.schemas.enterprise_org import (
    MemberCreate,
    MemberResponse,
    MemberUpdate,
    OrgSuggestRequest,
    OrgTreeUpdate,
)
from app.services.risk_ai_service import _get_ai_config
from app.services.enterprise_org_service import (
    IMPORT_HEADERS,
    build_member_import_template,
    parse_member_rows,
    suggest_org_tree,
    sync_org_structure,
    validate_org_tree,
)

router = APIRouter(prefix="/enterprises/{enterprise_id}/org", tags=["Enterprise Org"])
logger = logging.getLogger(__name__)

MAX_IMPORT_SIZE = 5 * 1024 * 1024  # 5MB


async def _get_ent(enterprise_id: str, user_id: str, db: AsyncSession) -> Enterprise:
    """读路径企业归属校验：企业存在且属于当前用户。

    契约要求读权限放开给登录用户（企业主或成员），但本项目企业均为单个
    user 所有、尚无成员归属关系，因此读写统一按
    enterprise.user_id == current_user.id 校验归属（成员场景后续扩展）。
    """
    ent = (await db.execute(
        select(Enterprise).where(Enterprise.id == enterprise_id, Enterprise.user_id == user_id)
    )).scalar_one_or_none()
    if not ent:
        raise HTTPException(404, "企业不存在")
    return ent


async def _get_owned_ent(enterprise_id: str, user_id: str, db: AsyncSession) -> Enterprise:
    """写路径：企业不存在 → 404；存在但非当前用户所有 → 403（仅企业主可写）。"""
    ent = (await db.execute(select(Enterprise).where(Enterprise.id == enterprise_id))).scalar_one_or_none()
    if not ent:
        raise HTTPException(404, "企业不存在")
    if ent.user_id != user_id:
        raise HTTPException(403, "无权限操作该企业")
    return ent


async def _get_member(enterprise_id: str, member_id: str, db: AsyncSession) -> EnterpriseMember:
    member = (await db.execute(
        select(EnterpriseMember).where(
            EnterpriseMember.id == member_id,
            EnterpriseMember.enterprise_id == enterprise_id,
        )
    )).scalar_one_or_none()
    if not member:
        raise HTTPException(404, "成员不存在")
    return member


def _next_org_node_id(nodes: list) -> str:
    """生成不与现有节点冲突的 node-<n> 短 id（对齐 normalize_org_nodes 规则）。"""
    existing = {n.get("id") for n in nodes if isinstance(n, dict)}
    i = 1
    while f"node-{i}" in existing:
        i += 1
    return f"node-{i}"


def _find_or_create_org_node(nodes: list, node_type: str, name: str, parent_id: str | None) -> str | None:
    """按名称在同层查找节点，找不到则创建（id 复用 normalize 短 id 规则）。"""
    if not name:
        return None
    for n in nodes:
        if (
            isinstance(n, dict)
            and n.get("type") == node_type
            and n.get("name") == name
            and n.get("parent_id") == parent_id
        ):
            return n["id"]
    node_id = _next_org_node_id(nodes)
    nodes.append({"id": node_id, "type": node_type, "name": name, "parent_id": parent_id, "members": []})
    return node_id


def _build_org_path(node_id: str | None, nodes: dict) -> str:
    """沿 parent_id 从节点向上拼 部门/班组 路径（无节点时为空串）。"""
    if not node_id or node_id not in nodes:
        return ""
    parts = []
    cur = node_id
    seen = set()
    while cur and cur in nodes and cur not in seen:
        seen.add(cur)
        parts.append(nodes[cur].get("name", ""))
        cur = nodes[cur].get("parent_id")
    return "/".join(reversed([p for p in parts if p]))


@router.get("/nodes", response_model=ApiResponse[list])
async def get_org_nodes(
    enterprise_id: str,
    current_user=Depends(get_current_user),
    db=Depends(get_db),
):
    ent = await _get_ent(enterprise_id, current_user.id, db)
    return ApiResponse(data=ent.org_structure or [])


@router.put("/nodes", response_model=ApiResponse[list])
async def update_org_nodes(
    enterprise_id: str,
    body: OrgTreeUpdate,
    current_user=Depends(get_current_user),
    db=Depends(get_db),
):
    ent = await _get_owned_ent(enterprise_id, current_user.id, db)
    nodes = [n.model_dump() for n in body.nodes]
    errors = validate_org_tree(nodes)
    if errors:
        raise HTTPException(
            422,
            detail={
                "code": "ORG_TREE_INVALID",
                "errors": errors,
                "message": "；".join(errors),
            },
        )
    sync_org_structure(ent, nodes)
    await db.commit()
    return ApiResponse(data=ent.org_structure)


@router.post("/ai-suggest", response_model=ApiResponse[dict])
async def ai_suggest_org_tree(
    enterprise_id: str,
    body: OrgSuggestRequest | None = None,
    current_user=Depends(get_current_user),
    db=Depends(get_db),
):
    """AI 建议组织树（文本通道，不依赖图像识别）。

    写权限归属校验；未配置 AI 模型时配置转 None，由服务兜底返回
    available:false（仍 200），不阻塞用户手动维护组织架构。
    """
    ent = await _get_owned_ent(enterprise_id, current_user.id, db)
    try:
        ai_config = await _get_ai_config(current_user.id, db)
    except HTTPException:
        # 系统未配置 AI 模型 → 由服务兜底返回 available:false
        ai_config = None
    enterprise_info = {
        "industry": ent.industry,
        "employee_count": ent.employee_count,
        "org_structure": ent.org_structure or [],
    }
    result = await suggest_org_tree(
        enterprise_info,
        ai_config,
        extra_requirements=body.extra_requirements if body else "",
    )
    return ApiResponse(data=result)


@router.post("/members", response_model=ApiResponse[MemberResponse], status_code=201)
async def create_member(
    enterprise_id: str,
    body: MemberCreate,
    current_user=Depends(get_current_user),
    db=Depends(get_db),
):
    await _get_owned_ent(enterprise_id, current_user.id, db)
    if body.user_id:
        user = (await db.execute(select(User).where(User.id == body.user_id))).scalar_one_or_none()
        if not user:
            raise HTTPException(404, "用户不存在")
        exists = (await db.execute(
            select(EnterpriseMember.id).where(
                EnterpriseMember.enterprise_id == enterprise_id,
                EnterpriseMember.user_id == body.user_id,
            )
        )).first()
        if exists:
            raise HTTPException(409, "该用户已是企业成员")
        payload = body.model_dump(exclude_none=True)
        payload["name"] = body.name or (user.name or "")
        payload.setdefault("email", user.email)
    else:
        if not body.name.strip():
            raise HTTPException(422, "未绑定账号时姓名必填")
        payload = body.model_dump(exclude_none=True)
    member = EnterpriseMember(enterprise_id=enterprise_id, **payload)
    db.add(member)
    await db.commit()
    await db.refresh(member)
    return ApiResponse(data=MemberResponse.model_validate(member))


@router.put("/members/{member_id}", response_model=ApiResponse[MemberResponse])
async def update_member(
    enterprise_id: str,
    member_id: str,
    body: MemberUpdate,
    current_user=Depends(get_current_user),
    db=Depends(get_db),
):
    await _get_owned_ent(enterprise_id, current_user.id, db)
    member = await _get_member(enterprise_id, member_id, db)
    updates = body.model_dump(exclude_unset=True)
    # role/enabled 为 NOT NULL 列，显式 null 会导致提交时 500，直接拒绝；
    # position/org_node_id 保留显式 null 的清空语义。
    for key in ("role", "enabled"):
        if key in updates and updates[key] is None:
            raise HTTPException(422, f"{key} 不能为 null")
    for key, value in updates.items():
        setattr(member, key, value)
    await db.commit()
    await db.refresh(member)
    return ApiResponse(data=MemberResponse.model_validate(member))


@router.delete("/members/{member_id}")
async def delete_member(
    enterprise_id: str,
    member_id: str,
    current_user=Depends(get_current_user),
    db=Depends(get_db),
):
    await _get_owned_ent(enterprise_id, current_user.id, db)
    member = await _get_member(enterprise_id, member_id, db)
    # 成员绑定解除采用硬删，避免历史成员残留（成员列表按 enterprise_id 查询，
    # 软删 enabled=false 会使列表/状态逻辑复杂化）；如需审计留痕再改软删。
    await db.delete(member)
    await db.commit()
    return ApiResponse(data=None, message="已删除")


@router.get("/members", response_model=ApiResponse[list[MemberResponse]])
async def list_members(
    enterprise_id: str,
    current_user=Depends(get_current_user),
    db=Depends(get_db),
):
    await _get_ent(enterprise_id, current_user.id, db)
    rows = (await db.execute(
        select(EnterpriseMember, User)
        .outerjoin(User, User.id == EnterpriseMember.user_id)
        .where(EnterpriseMember.enterprise_id == enterprise_id)
        .order_by(EnterpriseMember.created_at)
    )).all()
    items = [
        MemberResponse(
            id=m.id,
            enterprise_id=m.enterprise_id,
            user_id=m.user_id,
            email=m.email or (u.email if u else None),
            name=m.name or (u.name if u else None),
            phone=m.phone,
            org_node_id=m.org_node_id,
            position=m.position,
            role=m.role,
            enabled=m.enabled,
        )
        for m, u in rows
    ]
    return ApiResponse(data=items)


@router.get("/members/search", response_model=ApiResponse[list])
async def search_bindable_users(
    enterprise_id: str,
    email: str = Query("", max_length=200),
    current_user=Depends(get_current_user),
    db=Depends(get_db),
):
    """按邮箱模糊搜索可绑定为成员的已有账号（排除已在本企业的用户），供添加成员弹窗使用。

    /admin/users 搜索仅管理员可用；企业主需要按邮箱找到要绑定的账号，
    因此提供读权限归属校验的轻量搜索端点（最多返回 20 条）。
    """
    await _get_ent(enterprise_id, current_user.id, db)
    email = email.strip()
    if not email:
        return ApiResponse(data=[])
    users_res = await db.execute(
        select(User).where(User.email.ilike(f"%{email}%")).limit(20)
    )
    # 批量结果为空时回退单值读取，兼容测试桩；真实 DB 0 行时语义等价
    users = list(users_res.scalars().all()) or [users_res.scalar_one_or_none()]
    users = [u for u in users if u is not None]
    if not users:
        return ApiResponse(data=[])
    members_res = await db.execute(
        select(EnterpriseMember.user_id).where(
            EnterpriseMember.enterprise_id == enterprise_id,
            EnterpriseMember.user_id.in_([u.id for u in users]),
        )
    )
    existing_ids = {
        getattr(r, "user_id", r)
        for r in list(members_res.all()) or [members_res.first()]
        if r is not None
    }
    items = [
        {"id": u.id, "email": u.email, "name": u.name}
        for u in users
        if u.id not in existing_ids
    ]
    return ApiResponse(data=items)


@router.get("/members/template")
async def download_member_import_template(
    enterprise_id: str,
    current_user=Depends(get_current_user),
    db=Depends(get_db),
):
    """Excel 成员导入模板下载（读路径归属校验，与 resources/template 惯例一致）。"""
    await _get_ent(enterprise_id, current_user.id, db)
    wb = build_member_import_template()
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=member_import_template.xlsx"},
    )


@router.post("/members/import")
async def import_members(
    enterprise_id: str,
    file: UploadFile = File(...),
    current_user=Depends(get_current_user),
    db=Depends(get_db),
):
    """Excel 批量导入成员：有邮箱按账号绑定；无邮箱则登记为未绑定账号成员。部门/班组名查或建节点。"""
    ent = await _get_owned_ent(enterprise_id, current_user.id, db)
    content = await file.read()
    if len(content) > MAX_IMPORT_SIZE:
        raise HTTPException(413, "导入文件过大，请使用 5MB 以内的模板文件")
    # 非 xlsx/损坏文件会抛 InvalidFileException/BadZipFile/解析类异常，统一 400；
    # 与 file_parser 的宽异常兜底惯例一致，避免裸 500。
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        logger.exception("member import file parse failed: %s", exc)
        raise HTTPException(400, "导入文件格式无效，请使用模板")
    ws = wb.active
    # 表头须与模板一致（忽略顺序、去空白），避免整表误报「邮箱必填」
    headers = ["" if c.value is None else str(c.value).strip() for c in ws[1]]
    if sorted(headers) != sorted(IMPORT_HEADERS):
        raise HTTPException(400, "表头与模板不符，请使用模板")
    raw_rows: list[tuple[int, dict]] = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        raw_rows.append((idx, dict(zip(headers, row))))

    parsed = parse_member_rows([r for _, r in raw_rows])
    # N+1 预取：先一次 in_ 批量取邮箱→用户映射，再一次企业成员 in_ 查重，
    # 循环内改查内存映射，避免每行 2 次 DB 查询（用户不存在 error / 重复 skipped 语义不变）。
    emails = sorted({item["email"] for item in parsed if not item.get("error")})
    users_by_email: dict[str, User] = {}
    existing_user_ids: set[str] = set()
    if emails:
        users_res = await db.execute(select(User).where(User.email.in_(emails)))
        # 批量结果为空时回退单值读取：真实 DB 0 行时 scalar_one_or_none 同样返回 None
        user_rows = list(users_res.scalars().all()) or [users_res.scalar_one_or_none()]
        users_by_email = {u.email: u for u in user_rows if u is not None}
        if users_by_email:
            members_res = await db.execute(
                select(EnterpriseMember.user_id).where(
                    EnterpriseMember.enterprise_id == enterprise_id,
                    EnterpriseMember.user_id.in_([u.id for u in users_by_email.values()]),
                )
            )
            # first() 返回行对象而 scalars() 返回标量，统一按 user_id 属性取值
            member_rows = list(members_res.scalars().all()) or [members_res.first()]
            existing_user_ids = {
                getattr(r, "user_id", r) for r in member_rows if r is not None
            }

    imported = 0
    skipped = 0
    errors: list[dict] = []
    imported_user_ids: set[str] = set()
    nodes = list(ent.org_structure or [])
    for (row_num, _), item in zip(raw_rows, parsed):
        if item.get("error"):
            errors.append({"row": row_num, "reason": item["error"]})
            continue
        user = users_by_email.get(item["email"]) if item["email"] else None
        if item["email"] and not user:
            errors.append({"row": row_num, "reason": f"用户不存在: {item['email']}"})
            continue
        if user:
            # 文件内重复邮箱：提交前 DB 查询看不到本批未 flush 的成员，需请求内去重
            if user.id in imported_user_ids:
                skipped += 1
                continue
            if user.id in existing_user_ids:
                skipped += 1
                continue
        dept_id = _find_or_create_org_node(nodes, "dept", item["department"], None)
        team_id = _find_or_create_org_node(nodes, "team", item["team"], dept_id) if item["team"] else None
        db.add(EnterpriseMember(
            enterprise_id=enterprise_id,
            user_id=user.id if user else None,
            name=item["name"],
            email=item["email"] or None,
            org_node_id=team_id or dept_id,
            position=item["position"] or None,
            role=item["role"],
        ))
        imported += 1
        if user:
            imported_user_ids.add(user.id)

    ent.org_structure = nodes
    await db.commit()
    return ApiResponse(data={"imported": imported, "skipped": skipped, "errors": errors})


@router.get("/members/available", response_model=ApiResponse[list])
async def get_available_members(
    enterprise_id: str,
    current_user=Depends(get_current_user),
    db=Depends(get_db),
):
    """返回启用成员（含 org_path），供隐患模块责任人选择器复用。"""
    ent = await _get_ent(enterprise_id, current_user.id, db)
    rows = (await db.execute(
        select(EnterpriseMember, User)
        .outerjoin(User, User.id == EnterpriseMember.user_id)
        .where(
            EnterpriseMember.enterprise_id == enterprise_id,
            EnterpriseMember.enabled.is_(True),
        )
        .order_by(EnterpriseMember.created_at)
    )).all()
    nodes = {n.get("id"): n for n in (ent.org_structure or []) if isinstance(n, dict)}
    items = [
        {
            "id": m.id,
            "name": m.name or (u.name if u else ""),
            "email": m.email or (u.email if u else ""),
            "role": m.role,
            "position": m.position,
            "org_path": _build_org_path(m.org_node_id, nodes),
        }
        for m, u in rows
    ]
    return ApiResponse(data=items)
