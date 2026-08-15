"""隐患排查治理任务 11 测试：驾驶舱统计 + 台账/监管上报导出 + 未读数 + 权限。

测试风格与 tests/test_hazard_grade_api.py / test_risk_control_list.py 一致：
无 db fixture，端点用 FastAPI TestClient + dependency_overrides + SQL 文本分发
mock；统计口径直接测路由模块的纯函数 `_dashboard_payload`（today 可注入），
导出服务纯函数直接构造 ORM 记录验证 sheet/字段。

覆盖：
- 统计口径：整改及时率公式（应闭环/按期闭环）、平均整改周期（closed_at -
  created_at）、月度环比、未闭环/风险点、重大挂牌、超期（记录+任务）、
  扫码待确认
- 图表：类型分布/月度趋势（近 12 月窗口）/重大专表/企业对比（同账号多企业）
- 未读数：全企业 total / 当前用户 mine / by_type 分组
- 导出：台账 3 sheet 存在、监管字段白名单、脱敏字段（责任人姓名/联系方式/
  照片）不出现、责任单位 org 推导与缺省「—」、整改进度 content 优先
- 权限：非归属 404（dashboard 与两个导出端点）
"""

import io
from collections import namedtuple
from datetime import date, datetime
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.database import get_db
from app.dependencies import get_current_user
from app.models.data_dict import DataDict
from app.models.enterprise import Enterprise
from app.models.enterprise_org import EnterpriseMember
from app.models.hazard_management import (
    HazardInspectionTask,
    HazardRecord,
    HazardRectification,
)
from app.models.user import User
from app.routers import hazard_management
from app.services.data_dict_service import invalidate_dict_cache
from app.services.hazard_export_service import (
    LEDGER_HEADERS,
    REPORT_HEADERS,
    build_ledger_workbook,
    build_report_workbook,
    resolve_department_name,
)


_UserRow = namedtuple("_UserRow", "id name")


# ── mock 工具 ──

def _scalar(value):
    res = MagicMock()
    res.scalar_one_or_none.return_value = value
    return res


def _scalars(values):
    res = MagicMock()
    scalars_mock = MagicMock()
    scalars_mock.all.return_value = list(values)
    res.scalars.return_value = scalars_mock
    return res


def _rows(values):
    """两列查询（select 多列）结果 mock：`.all()` 与 `.scalars().all()` 双兼容。"""
    res = MagicMock()
    res.all.return_value = list(values)
    scalars_mock = MagicMock()
    scalars_mock.all.return_value = list(values)
    res.scalars.return_value = scalars_mock
    return res


def _first(value):
    res = MagicMock()
    res.first.return_value = value
    return res


def _ent(**kw):
    e = Enterprise(id="e1", user_id="u1", name="甲公司", hazard_closure_mode="standard")
    for k, v in kw.items():
        setattr(e, k, v)
    return e


def _record(**kw):
    r = HazardRecord(
        id=kw.pop("id", "r1"),
        enterprise_id=kw.pop("enterprise_id", "e1"),
        code=kw.pop("code", "HD-001"),
        source_type=kw.pop("source_type", "report"),
        title=kw.pop("title", "配电箱门破损"),
        description=kw.pop("description", "配电箱门变形无法闭合"),
        status=kw.pop("status", "registered"),
    )
    for k, v in kw.items():
        setattr(r, k, v)
    return r


def _task(**kw):
    t = HazardInspectionTask(
        id=kw.pop("id", "t1"),
        plan_id=kw.pop("plan_id", "p1"),
        enterprise_id=kw.pop("enterprise_id", "e1"),
        title=kw.pop("title", "日常排查"),
        status=kw.pop("status", "pending"),
        due_at=kw.pop("due_at", datetime(2026, 8, 1, 18, 0)),
    )
    for k, v in kw.items():
        setattr(t, k, v)
    return t


def _dict_row(dict_type, code, label):
    return DataDict(dict_type=dict_type, code=code, label=label,
                    value={}, scope="system", sort_order=1, enabled=True,
                    is_system=True)


def _status_rows(labels=None):
    labels = labels or {
        "registered": "已登记", "rectifying": "整改中", "closed": "已闭环",
    }
    return [_dict_row("record_status_label", k, v) for k, v in labels.items()]


def _dashboard_db(
    ent,
    *,
    records=None,
    tasks=None,
    approved_ids=None,
    unread_rows=None,
    owned_rows=None,
    status_rows=None,
    member=None,
):
    """按 SQL 文本特征分发（参照 test_hazard_grade_api._grade_db）。

    - enterprises：id 查询 → 企业；user_id 查询 → 同账号名下企业
    - hazard_approvals：审批历史 record_id 集合
    - hazard_records：单企业全量 ORM；IN 查询 → (enterprise_id, status) 行
    - hazard_inspection_tasks / hazard_notifications 各一查询
    """
    records = records or []
    tasks = tasks or []
    approved_ids = approved_ids or []
    unread_rows = unread_rows or []
    owned_rows = owned_rows or []
    status_rows = status_rows or []
    db = AsyncMock()
    db.added = []
    db.add = MagicMock(side_effect=lambda obj: db.added.append(obj))

    def fake_execute(stmt, *params):
        text = str(stmt)
        if "FROM enterprises" in text:
            if "enterprises.user_id =" in text:  # WHERE 名下企业查询（区分 SELECT * 的企业查询）
                return _rows(owned_rows)
            return _scalar(ent)
        if "FROM enterprise_members" in text:
            return _first(member)
        if "FROM hazard_approvals" in text:
            return _scalars(approved_ids)
        if "FROM hazard_records" in text:
            if "hazard_records.enterprise_id IN" in text:
                return _rows(status_rows)
            return _scalars(records)
        if "FROM hazard_inspection_tasks" in text:
            return _scalars(tasks)
        if "FROM hazard_notifications" in text:
            return _rows(unread_rows)
        return _scalars([])

    db.execute.side_effect = fake_execute
    return db


@pytest.fixture()
def client():
    app = FastAPI()
    app.include_router(hazard_management.router)

    def _override_user():
        return User(id="u1", email="a@b.c", name="A", role="user")

    app.dependency_overrides[get_current_user] = _override_user
    app.dependency_overrides[get_db] = lambda: _dashboard_db(_ent())
    with TestClient(app) as test_client:
        yield test_client


@pytest.fixture(autouse=True)
def _clear_dict_cache():
    """每个测试结束后清空数据字典进程内缓存，避免污染同进程其他测试模块。"""
    yield
    invalidate_dict_cache()


# ── 统计口径：整改及时率公式 / 平均整改周期 / 月度环比（纯函数，today 注入） ──

def test_rectification_rate_formula_and_cycle_days():
    today = date(2026, 8, 15)
    recs = [
        _record(id="r1", code="HD-001", status="closed",
                deadline=date(2026, 8, 5),
                closed_at=datetime(2026, 8, 3, 10, 0),
                created_at=datetime(2026, 8, 1, 9, 0)),   # 按期：closed_at <= deadline
        _record(id="r2", code="HD-002", status="closed",
                deadline=date(2026, 8, 10),
                closed_at=datetime(2026, 8, 12, 10, 0),   # 超期闭环：closed_at > deadline
                created_at=datetime(2026, 8, 2, 9, 0)),
        _record(id="r3", code="HD-003", status="rectifying",
                deadline=date(2026, 8, 14)),              # 本月超期（deadline < today）
        _record(id="r4", code="HD-004", status="rectifying",
                deadline=date(2026, 8, 20)),              # 本月内但未超期：不算应闭环
        _record(id="r5", code="HD-005", status="closed",
                deadline=date(2026, 7, 31)),              # 上月 deadline：不算本月应闭环
    ]
    payload = hazard_management._dashboard_payload(
        recs, [], set(), [], [], [], "u1", today=today)
    m = payload["metrics"]
    assert m["due_this_month"] == 3   # r1/r2（closed）+ r3（本月超期 rectifying）
    assert m["on_time_closed"] == 1   # 仅 r1 按期
    assert m["rectification_rate"] == round(1 / 3 * 100, 1)
    # 平均整改周期：本月闭环 r1（2 天）+ r2（10 天）
    assert m["avg_rectification_days"] == 6.0


def test_rectification_rate_no_due_records_is_none():
    today = date(2026, 8, 15)
    recs = [_record(status="registered", deadline=None)]
    payload = hazard_management._dashboard_payload(
        recs, [], set(), [], [], [], "u1", today=today)
    assert payload["metrics"]["due_this_month"] == 0
    assert payload["metrics"]["rectification_rate"] is None
    assert payload["metrics"]["avg_rectification_days"] is None


def test_monthly_new_mom_formula():
    today = date(2026, 8, 15)
    recs = [
        _record(id="r1", created_at=datetime(2026, 8, 1, 9, 0)),
        _record(id="r2", created_at=datetime(2026, 8, 3, 9, 0)),
        _record(id="r3", created_at=datetime(2026, 7, 10, 9, 0)),   # 上月 1 条
    ]
    payload = hazard_management._dashboard_payload(
        recs, [], set(), [], [], [], "u1", today=today)
    m = payload["metrics"]
    assert m["monthly_new"] == 2
    assert m["monthly_new_mom"] == 100.0   # (2-1)/1*100


def test_monthly_new_mom_zero_last_month_is_none():
    today = date(2026, 8, 15)
    recs = [_record(id="r1", created_at=datetime(2026, 8, 1, 9, 0))]
    payload = hazard_management._dashboard_payload(
        recs, [], set(), [], [], [], "u1", today=today)
    assert payload["metrics"]["monthly_new"] == 1
    assert payload["metrics"]["monthly_new_mom"] is None


# ── 驾驶舱端点：指标/图表/未读/企业对比 ──

def test_dashboard_metrics_and_charts(client):
    today = date(2026, 8, 15)
    recs = [
        _record(id="r1", code="HD-001", status="rectifying", level="major",
                hazard_type="fire", object_id="o1",
                created_at=datetime(2026, 8, 1, 9, 0),
                deadline=date(2026, 8, 10), rectification_user_id="u2"),
        _record(id="r2", code="HD-002", status="registered", level=None,
                hazard_type="fire", object_id="o1",
                source_type="report",
                created_at=datetime(2026, 8, 2, 9, 0)),
        _record(id="r3", code="HD-003", status="closed", level="major",
                hazard_type="equipment", object_id="o2",
                created_at=datetime(2026, 7, 5, 9, 0),
                deadline=date(2026, 7, 30),
                closed_at=datetime(2026, 7, 28, 10, 0)),
        _record(id="r4", code="HD-004", status="rectifying", level="general",
                hazard_type=None, object_id=None,
                created_at=datetime(2026, 8, 3, 9, 0),
                deadline=date(2026, 8, 20)),
    ]
    tasks = [_task(status="overdue"), _task(id="t2", status="pending")]
    unread = [("u1", "overdue"), ("u2", "overdue"), ("u1", "upcoming")]
    owned = [("e1", "甲公司"), ("e2", "乙公司")]
    status_rows = [
        ("e1", "rectifying"), ("e1", "registered"), ("e1", "closed"),
        ("e2", "rectifying"), ("e2", "closed"),
    ]
    db = _dashboard_db(
        _ent(), records=recs, tasks=tasks,
        approved_ids=["r3"], unread_rows=unread,
        owned_rows=owned, status_rows=status_rows,
    )
    client.app.dependency_overrides[get_db] = lambda: db
    with patch("app.routers.hazard_management._today", return_value=today):
        resp = client.get("/enterprises/e1/hazard-inspection/dashboard")
    assert resp.status_code == 200
    assert resp.json()["code"] == 0
    data = resp.json()["data"]
    m = data["metrics"]
    assert m["open_hazards"] == 3          # r1/r2/r4（closed 只有 r3）
    assert m["open_risk_points"] == 1      # 未闭环记录去重 object_id：仅 o1（r3 已闭环不计）
    assert m["major_count"] == 1           # 当前 major 未闭环：r1
    assert m["major_approved"] == 1        # r3 有 approve 历史
    assert m["overdue_records"] == 1       # r1 rectifying 且 deadline < today
    assert m["overdue_tasks"] == 1         # t1 overdue
    assert m["overdue_count"] == 2
    assert m["monthly_new"] == 3           # r1/r2/r4 本月新增
    assert m["scan_pending"] == 1          # r2 report/registered
    # 类型分布：fire 2、equipment 1、未分类 1
    dist = data["charts"]["type_distribution"]
    assert dist == [
        {"hazard_type": "fire", "count": 2},
        {"hazard_type": "equipment", "count": 1},
        {"hazard_type": "未分类", "count": 1},
    ]
    # 月度趋势：近 12 月窗口，2026-07=1、2026-08=3，其余 0
    trend = data["charts"]["monthly_trend"]
    assert len(trend) == 12
    assert trend[0]["month"] == "2025-09"
    assert trend[-2]["month"] == "2026-07" and trend[-2]["count"] == 1
    assert trend[-1]["month"] == "2026-08" and trend[-1]["count"] == 3
    # 重大专表：deadline 升序（r3 07-30 早于 r1 08-10），字段白名单 code/title/deadline/status
    majors = data["charts"]["major_records"]
    assert [x["code"] for x in majors] == ["HD-003", "HD-001"]
    assert set(majors[0].keys()) == {"code", "title", "deadline", "status"}
    # 企业对比：e1=2 未闭环（r1/r2/r4）、e2=1，按未闭环降序
    comp = data["charts"]["enterprise_comparison"]
    assert comp == [
        {"enterprise_id": "e1", "name": "甲公司", "open_count": 2},
        {"enterprise_id": "e2", "name": "乙公司", "open_count": 1},
    ]


def test_dashboard_unread_total_mine_by_type(client):
    unread = [("u1", "overdue"), ("u2", "overdue"), ("u1", "upcoming"),
              ("u3", "review_due")]
    db = _dashboard_db(_ent(), unread_rows=unread)
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.get("/enterprises/e1/hazard-inspection/dashboard")
    assert resp.status_code == 200
    unread_data = resp.json()["data"]["unread"]
    assert unread_data["total"] == 4      # 本企业全部未读
    assert unread_data["mine"] == 2       # 当前用户 u1 未读
    assert unread_data["by_type"] == {"overdue": 2, "upcoming": 1, "review_due": 1}


def test_dashboard_major_approved_includes_pending_approval(client):
    recs = [_record(id="r1", status="pending_approval", level="major"),
            _record(id="r2", status="rectifying", level="general")]
    db = _dashboard_db(_ent(), records=recs, approved_ids=["r9"])
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.get("/enterprises/e1/hazard-inspection/dashboard")
    assert resp.status_code == 200
    m = resp.json()["data"]["metrics"]
    assert m["major_approved"] == 1       # r1 当前 pending_approval（无审批历史也计入）


def test_dashboard_not_owned_404(client):
    ent = _ent(user_id="u2")
    db = _dashboard_db(ent, member=None)
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.get("/enterprises/e1/hazard-inspection/dashboard")
    assert resp.status_code == 404
    assert "企业不存在" in resp.json()["detail"]


# ── 导出服务纯函数：台账 3 sheet / 监管白名单 / 脱敏 / 责任单位 ──

def test_build_ledger_workbook_three_sheets_and_overdue_filter():
    today = date(2026, 8, 15)
    recs = [
        _record(id="r1", code="HD-001", status="rectifying", level="major",
                deadline=date(2026, 8, 10),
                rectification_user_id="u2",
                photo_urls=["a.jpg", "b.jpg"],
                rectification_plan={"goal": "更换箱门"},
                object_id="o1", measure_id="m1",
                hazard_type="fire", source_type="report",
                created_at=datetime(2026, 8, 1, 9, 0)),
        _record(id="r2", code="HD-002", status="rectifying", level="general",
                deadline=date(2026, 8, 20),   # 未超期：不进 sheet2
                rectification_user_id=None),
    ]
    wb = build_ledger_workbook(
        recs, today=today,
        object_names={"o1": "1#配电房"},
        measure_names={"m1": "更换箱门并加锁"},
        user_names={"u2": "李四"},
        status_labels={"rectifying": "整改中"},
        source_labels={"report": "扫码上报"},
        hazard_type_labels={"fire": "火灾"},
    )
    assert wb.sheetnames == ["台账", "超期清单", "重大隐患"]
    ws = wb["台账"]
    assert [c.value for c in ws[1]] == LEDGER_HEADERS
    assert ws.max_row == 3
    row = [c.value for c in ws[2]]
    assert row[0] == "HD-001"
    assert row[3] == "火灾"               # hazard_type 标签
    assert row[6] == "整改中"             # status 标签
    assert row[7] == "扫码上报"           # source_type 标签
    assert row[8] == "1#配电房"           # 风险点名称解析
    assert row[9] == "更换箱门并加锁"      # 管控措施描述解析
    assert row[11] == "a.jpg、b.jpg"      # 照片拼接
    assert "更换箱门" in row[12]          # 治理方案 JSON
    assert row[14] == "李四"              # 整改责任人姓名（台账含敏感字段）
    ws2 = wb["超期清单"]
    assert [c.value for c in ws2[1]] == ["编号", "标题", "等级", "状态",
                                         "整改期限", "整改责任人", "超期天数"]
    assert ws2.max_row == 2               # 仅 r1 超期
    assert ws2["G2"].value == 5           # 超期天数 = 2026-08-15 - 08-10
    ws3 = wb["重大隐患"]
    assert [c.value for c in ws3[1]] == ["编号", "标题", "等级", "状态",
                                         "整改期限", "判定依据", "整改责任人", "登记时间"]
    assert ws3.max_row == 2               # 仅 r1 major


def test_build_ledger_workbook_names_fallback_to_ids():
    recs = [_record(id="r1", status="rectifying", level="general",
                    deadline=date(2026, 8, 10), object_id="o9",
                    rectification_user_id="u9")]
    wb = build_ledger_workbook(recs, today=date(2026, 8, 15))
    ws = wb["台账"]
    row = [c.value for c in ws[2]]
    assert row[8] == "o9"    # 名称映射未命中 → 回退原始 id
    assert row[14] == "u9"


def test_resolve_department_name_walks_to_dept_and_defaults():
    node_map = {
        "d1": {"id": "d1", "type": "dept", "name": "生产部", "parent_id": None},
        "t1": {"id": "t1", "type": "team", "name": "机修班", "parent_id": "d1"},
        "p1": {"id": "p1", "type": "position", "name": "班组长", "parent_id": "t1"},
        "iso": {"id": "iso", "type": "team", "name": "独立班组", "parent_id": None},
    }
    assert resolve_department_name("p1", node_map) == "生产部"  # 岗位→班组→部门
    assert resolve_department_name("d1", node_map) == "生产部"  # 本身就是部门
    assert resolve_department_name("iso", node_map) == "—"       # 无部门祖先
    assert resolve_department_name("missing", node_map) == "—"
    assert resolve_department_name(None, node_map) == "—"


def test_build_report_workbook_whitelist_and_desensitization():
    recs = [
        _record(id="r1", code="HD-001", title="储罐超压", location="罐区",
                level="major", grading_basis="符合危化品判定要点",
                deadline=date(2026, 8, 30), rectification_user_id="u2",
                created_at=datetime(2026, 8, 1, 9, 0)),
        _record(id="r2", code="HD-002", title="配电箱门破损", location=None,
                level="general", grading_basis=None, deadline=None,
                rectification_user_id="u9",  # 无部门映射 → 缺省「—」
                created_at=datetime(2026, 8, 2, 9, 0)),
    ]
    wb = build_report_workbook(
        recs,
        org_dept_map={"u2": "生产部"},
        progress_map={"r1": "已更换安全阀", "r2": "整改中"},
    )
    ws = wb.active
    assert ws.title == "监管上报台账"
    headers = [c.value for c in ws[1]]
    assert headers == REPORT_HEADERS
    assert headers == ["编号", "名称", "位置", "等级", "判定依据", "整改期限",
                       "责任单位", "整改进度"]
    # 脱敏：白名单列名不出现责任人姓名/联系方式/照片
    assert "责任人" not in headers
    assert "电话" not in headers
    assert "照片" not in headers
    assert ws.max_row == 3
    # 创建时间倒序：r2（08-02）在前、r1（08-01）在后
    row2 = [c.value for c in ws[2]]
    assert row2[0] == "HD-002"
    assert row2[1] == "配电箱门破损"
    assert row2[2] == "-"
    assert row2[4] == "-"
    assert row2[5] == "-"
    assert row2[6] == "—"        # 无部门映射 → 缺省
    assert row2[7] == "整改中"
    row1 = [c.value for c in ws[3]]
    assert row1[0] == "HD-001"
    assert row1[1] == "储罐超压"
    assert row1[2] == "罐区"
    assert row1[4] == "符合危化品判定要点"
    assert row1[5] == "2026-08-30"
    assert row1[6] == "生产部"    # 责任单位 org 推导
    assert row1[7] == "已更换安全阀"  # 整改进度 content 优先


# ── 导出端点：台账 / 监管上报 / 权限 404 ──

def _export_db(
    ent,
    *,
    records=None,
    dict_rows=None,
    user_rows=None,
    object_rows=None,
    measure_rows=None,
    members=None,
    rect_rows=None,
):
    """导出端点 mock：覆盖 enterprises/data_dicts/users/risk_objects/risk_measures/
    hazard_rectifications/enterprise_members 查询。"""
    records = records or []
    dict_rows = dict_rows or (
        _status_rows()
        + [_dict_row("source_type", "report", "扫码上报")]
        + [_dict_row("hazard_type", "fire", "火灾")]
    )
    db = AsyncMock()

    def fake_execute(stmt, *params):
        text = str(stmt)
        if "FROM enterprises" in text:
            return _scalar(ent)
        if "FROM data_dicts" in text:
            return _scalars(dict_rows)
        if "FROM users" in text:
            return _rows(user_rows or [])
        if "FROM risk_objects" in text:
            return _rows(object_rows or [])
        if "FROM risk_measures" in text:
            return _rows(measure_rows or [])
        if "FROM hazard_rectifications" in text:
            return _rows(rect_rows or [])
        if "FROM enterprise_members" in text:
            if members:
                return _scalars(members)
            return _first(None)
        if "FROM hazard_records" in text:
            return _scalars(records)
        return _scalars([])

    db.execute.side_effect = fake_execute
    return db


def test_export_ledger_xlsx(client):
    rec = _record(id="r1", code="HD-001", status="rectifying", level="major",
                  deadline=date(2026, 8, 10), rectification_user_id="u2",
                  object_id="o1", measure_id="m1",
                  hazard_type="fire", source_type="report")
    db = _export_db(
        _ent(), records=[rec],
        user_rows=[_UserRow("u2", "李四")],
        object_rows=[_UserRow("o1", "1#配电房")],
        measure_rows=[_UserRow("m1", "更换箱门并加锁")],
    )
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.get("/enterprises/e1/hazard-inspection/export/ledger.xlsx")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert "hazard_ledger.xlsx" in resp.headers["content-disposition"]
    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["台账", "超期清单", "重大隐患"]
    ws = wb["台账"]
    assert ws["A1"].value == "编号"
    row = [c.value for c in ws[2]]
    assert row[0] == "HD-001"
    assert row[14] == "李四"
    assert row[8] == "1#配电房"


def test_export_ledger_xlsx_other_enterprise_404(client):
    db = _export_db(_ent(user_id="u2"))
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.get("/enterprises/e1/hazard-inspection/export/ledger.xlsx")
    assert resp.status_code == 404


def test_export_report_xlsx_whitelist_and_desensitized(client):
    ent = _ent()
    ent.org_structure = [
        {"id": "d1", "type": "dept", "name": "生产部", "parent_id": None},
        {"id": "t1", "type": "team", "name": "机修班", "parent_id": "d1"},
    ]
    rec = _record(id="r1", code="HD-001", title="储罐超压", location="罐区",
                  level="major", grading_basis="符合危化品判定要点",
                  deadline=date(2026, 8, 30), rectification_user_id="u2",
                  photo_urls=["sensitive.jpg"])
    member = EnterpriseMember(id="m1", enterprise_id="e1", user_id="u2",
                              org_node_id="t1", role="member", enabled=True)
    rect = HazardRectification(id="rc1", record_id="r1", content="已更换安全阀",
                               user_id="u2")
    db = _export_db(ent, records=[rec], members=[member], rect_rows=[rect])
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.get("/enterprises/e1/hazard-inspection/export/report.xlsx")
    assert resp.status_code == 200
    assert "hazard_report.xlsx" in resp.headers["content-disposition"]
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers == REPORT_HEADERS
    # 脱敏：不出现责任人姓名/联系方式/照片/登记人等敏感列
    assert "责任人" not in headers
    assert "电话" not in headers
    assert "照片" not in headers
    row = [c.value for c in ws[2]]
    assert row[6] == "生产部"          # 岗位→班组→部门 推导
    assert row[7] == "已更换安全阀"     # 最近整改 content 优先
    # 全部单元格不含照片 URL（脱敏落地校验）
    cells = [c.value for row_ws in ws.iter_rows() for c in row_ws]
    assert all("sensitive.jpg" not in str(v) for v in cells if v is not None)


def test_export_report_xlsx_progress_falls_back_to_status_label(client):
    rec = _record(id="r1", status="rectifying", level="general")
    db = _export_db(_ent(), records=[rec])
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.get("/enterprises/e1/hazard-inspection/export/report.xlsx")
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws["H2"].value == "整改中"   # 无整改记录 → 状态标签
    assert ws["G2"].value == "—"        # 无成员映射 → 缺省


def test_export_report_xlsx_other_enterprise_404(client):
    db = _export_db(_ent(user_id="u2"))
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.get("/enterprises/e1/hazard-inspection/export/report.xlsx")
    assert resp.status_code == 404


# ── 端点无鉴权 401 ──

def test_dashboard_requires_auth():
    app = FastAPI()
    app.include_router(hazard_management.router)
    app.dependency_overrides[get_db] = lambda: _dashboard_db(_ent())
    with TestClient(app) as test_client:
        resp = test_client.get("/enterprises/e1/hazard-inspection/dashboard")
    assert resp.status_code == 401


def test_export_ledger_requires_auth():
    app = FastAPI()
    app.include_router(hazard_management.router)
    app.dependency_overrides[get_db] = lambda: _export_db(_ent())
    with TestClient(app) as test_client:
        resp = test_client.get(
            "/enterprises/e1/hazard-inspection/export/ledger.xlsx")
    assert resp.status_code == 401
