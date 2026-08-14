"""风险分级管控清单 + Excel 导出 + 重大风险公示后端测试。

服务为纯函数（无 db fixture）；端点用独立 FastAPI 应用挂载 router，
dependency_overrides 替换鉴权与 DB 依赖（参考 test_risk_notice_card_api.py）。
"""
import io
from datetime import datetime
from unittest.mock import AsyncMock, MagicMock

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.database import get_db
from app.dependencies import get_current_user
from app.models.data_dict import DataDict
from app.models.enterprise import Enterprise, EnterpriseFloor
from app.models.risk_management import (
    RiskEvent,
    RiskMeasure,
    RiskObject,
    RiskUnit,
    RiskZone,
)
from app.routers import public_risk, risk_management
from app.services.risk_control_list_service import (
    build_ledger_workbook,
    default_control_level,
    desensitize,
    flatten_rows,
)


# ── 服务纯函数 ──

def test_default_control_level_from_dict():
    mapping = {"重大": "企业", "较大": "部门", "一般": "班组", "低": "岗位"}
    assert default_control_level(mapping, "重大") == "企业"
    assert default_control_level(mapping, None) == "岗位"
    assert default_control_level(mapping, "未知等级") == "岗位"


def test_build_ledger_workbook():
    rows = [
        {"zone": "储罐区", "object": "1#储罐", "unit": "阀门组",
         "accident": "泄漏", "inherent": "重大", "current": "一般",
         "control_level": "班组", "measures": "报警器年检", "unit_name": "生产部",
         "person": "李四", "phone": "13800000000"},
        {"zone": "罐区", "object": "2#储罐", "unit": "-",
         "accident": "火灾", "inherent": "一般", "current": "一般",
         "control_level": "企业", "measures": "-", "unit_name": "安全部",
         "person": "王五", "phone": "13900000000"},
    ]
    wb = build_ledger_workbook(rows)
    assert wb.sheetnames == ["风险管控清单", "等级层级汇总"]
    ws = wb.active
    assert ws.title == "风险管控清单"
    assert ws["A1"].value == "分区"
    assert ws.max_row == 3
    assert ws["B2"].value == "1#储罐"
    assert ws["G2"].value == "班组"
    assert ws["J2"].value == "李四"
    assert ws["A1"].font.bold is True
    # sheet2：固有等级/管控层级汇总，按固定顺序，数量正确
    ws2 = wb["等级层级汇总"]
    assert ws2["A1"].value == "固有等级"
    assert ws2["B1"].value == "数量"
    assert [ws2.cell(row=i, column=1).value for i in range(2, 6)] == ["低", "一般", "较大", "重大"]
    assert [ws2.cell(row=i, column=2).value for i in range(2, 6)] == [0, 1, 0, 1]
    assert ws2.cell(row=6, column=1).value is None  # 空行分隔
    assert ws2["A7"].value == "管控层级"
    assert ws2["B7"].value == "数量"
    assert [ws2.cell(row=i, column=1).value for i in range(8, 12)] == ["岗位", "班组", "部门", "企业"]
    assert [ws2.cell(row=i, column=2).value for i in range(8, 12)] == [0, 1, 0, 1]
    assert ws2["A1"].font.bold is True


def test_flatten_rows_includes_zone_id():
    zone = RiskZone(id="z1", enterprise_id="e1", floor_id="f1", name="储罐区")
    obj = RiskObject(id="o1", enterprise_id="e1", zone_id="z1", name="1#储罐")
    obj.events = [RiskEvent(accident_type="泄漏", risk_level="重大", inherent_risk_level="重大")]
    zone.objects = [obj]
    rows = flatten_rows([zone], {"重大": "企业"})
    assert len(rows) == 1
    assert rows[0]["zone_id"] == "z1"
    assert rows[0]["object_id"] == "o1"
    assert rows[0]["inherent"] == "重大"
    assert rows[0]["control_level"] == "企业"
    # 位置缺省回退 "-"
    assert rows[0]["location"] == "-"


def test_flatten_rows_unit_measures_and_mapping_fallback():
    zone = RiskZone(id="z1", enterprise_id="e1", floor_id="f1", name="储罐区")
    obj = RiskObject(id="o1", enterprise_id="e1", zone_id="z1", name="1#储罐",
                     location="储罐区东侧", responsible_unit="生产部",
                     responsible_person="李四", contact_phone="13800000000")
    unit = RiskUnit(id="u1", object_id="o1", name="阀门组")
    ev = RiskEvent(accident_type="泄漏", risk_level="较大", inherent_risk_level="重大")
    ev.measures = [RiskMeasure(event_id="m1", measure_category="工程技术", description="报警器年检")]
    unit.events = [ev]
    obj.units = [unit]
    zone.objects = [obj]
    rows = flatten_rows([zone], {"较大": "部门"})
    assert len(rows) == 1
    row = rows[0]
    assert row["unit"] == "阀门组"
    assert row["measures"] == "工程技术:报警器年检"
    assert row["control_level"] == "部门"
    assert row["unit_name"] == "生产部"
    assert row["person"] == "李四"
    assert row["phone"] == "13800000000"
    assert row["location"] == "储罐区东侧"


def test_desensitize_drops_person_phone_and_internal_keys():
    rows = [{"zone_id": "z1", "object_id": "o1", "zone": "储罐区", "object": "1#储罐",
             "unit": "-", "location": "储罐区东侧", "accident": "泄漏", "inherent": "重大", "current": "重大",
             "control_level": "企业", "measures": "报警器年检", "unit_name": "生产部",
             "person": "李四", "phone": "13800000000"}]
    out = desensitize(rows)
    assert len(out) == 1
    assert "person" not in out[0]
    assert "phone" not in out[0]
    assert "zone_id" not in out[0]
    assert "object_id" not in out[0]
    assert out[0]["zone"] == "储罐区"
    assert out[0]["object"] == "1#储罐"
    assert out[0]["unit_name"] == "生产部"
    assert out[0]["location"] == "储罐区东侧"


# ── 端点测试公共设施 ──

def _ent(**kw):
    e = Enterprise(id="e1", user_id="u1", name="甲公司")
    for k, v in kw.items():
        setattr(e, k, v)
    return e


def _scalar_result(value):
    res = MagicMock()
    res.scalar_one_or_none.return_value = value
    return res


def _rows_result(rows):
    res = MagicMock()
    res.scalars.return_value.all.return_value = rows
    return res


def _dict_row(code, label, value):
    return DataDict(dict_type="control_level_map", code=code, label=label,
                    value=value, scope="system", sort_order=1, enabled=True)


def _default_floor():
    return EnterpriseFloor(id="f1", enterprise_id="e1", name="默认总图",
                           sort_order=0, is_default=True)


def _db(ent, zones=None, dict_rows=None):
    """按 SQL 文本特征分发：enterprises→企业、enterprise_floors→默认楼层、
    data_dicts→层级映射、risk_zones→分区树。"""
    zones = zones or []
    dict_rows = dict_rows or [
        _dict_row("major", "重大→企业", {"level": "重大", "control_level": "企业"}),
        _dict_row("large", "较大→部门", {"level": "较大", "control_level": "部门"}),
        _dict_row("general", "一般→班组", {"level": "一般", "control_level": "班组"}),
        _dict_row("low", "低→岗位", {"level": "低", "control_level": "岗位"}),
    ]
    db = AsyncMock()
    db.add = MagicMock()

    def fake_execute(stmt, *params):
        text = str(stmt)
        if "FROM enterprises" in text:
            return _scalar_result(ent)
        if "FROM enterprise_floors" in text:
            if "enterprise_floors.id =" in text:
                # 显式 floor_id 查询：楼层不存在（如他企业楼层）→ None
                return _scalar_result(None)
            return _scalar_result(_default_floor())
        if "FROM data_dicts" in text:
            return _rows_result(dict_rows)
        if "FROM risk_zones" in text:
            return _rows_result(zones)
        return _rows_result([])

    db.execute.side_effect = fake_execute
    return db


def _zone_with_events():
    zone = RiskZone(id="z1", enterprise_id="e1", floor_id="f1", name="储罐区")
    obj = RiskObject(id="o1", enterprise_id="e1", zone_id="z1", name="1#储罐",
                     location="储罐区东侧",
                     responsible_unit="生产部", responsible_person="李四",
                     contact_phone="13800000000")
    obj.events = [
        RiskEvent(accident_type="泄漏", risk_level="重大", inherent_risk_level="重大",
                  control_level="企业"),
        RiskEvent(accident_type="火灾", risk_level="一般", inherent_risk_level="较大"),
    ]
    zone.objects = [obj]
    return zone


@pytest.fixture()
def client():
    from app.models.user import User

    app = FastAPI()
    app.include_router(risk_management.router)

    def _override_user():
        return User(id="u1", email="a@b.c", name="A", role="admin")

    app.dependency_overrides[get_current_user] = _override_user
    app.dependency_overrides[get_db] = lambda: _db(_ent())
    with TestClient(app) as test_client:
        yield test_client


@pytest.fixture()
def public_client():
    app = FastAPI()
    app.include_router(public_risk.router)
    app.dependency_overrides[get_db] = lambda: _db(None)
    with TestClient(app) as test_client:
        yield test_client


# ── control-list ──

def test_control_list_requires_auth():
    app = FastAPI()
    app.include_router(risk_management.router)
    app.dependency_overrides[get_db] = lambda: _db(_ent())
    with TestClient(app) as test_client:
        resp = test_client.get("/enterprises/e1/risk-management/control-list")
    assert resp.status_code == 401


def test_control_list_enterprise_not_found_404(client):
    client.app.dependency_overrides[get_db] = lambda: _db(None)
    resp = client.get("/enterprises/e1/risk-management/control-list")
    assert resp.status_code == 404
    assert "企业不存在" in resp.json()["detail"]


def test_control_list_pagination_and_filters(client):
    zone = _zone_with_events()
    db = _db(_ent(), zones=[zone])
    client.app.dependency_overrides[get_db] = lambda: db

    resp = client.get("/enterprises/e1/risk-management/control-list")
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert data["total"] == 2
    assert len(data["items"]) == 2
    # 内部键从响应去除
    assert "zone_id" not in data["items"][0]
    assert "object_id" not in data["items"][0]

    # level 筛选：匹配 current 或 inherent
    resp = client.get("/enterprises/e1/risk-management/control-list", params={"level": "重大"})
    data = resp.json()["data"]
    assert data["total"] == 1
    assert data["items"][0]["accident"] == "泄漏"

    # control_level 筛选
    resp = client.get("/enterprises/e1/risk-management/control-list",
                      params={"control_level": "企业"})
    assert resp.json()["data"]["total"] == 1

    # zone_id 筛选
    resp = client.get("/enterprises/e1/risk-management/control-list", params={"zone_id": "z1"})
    assert resp.json()["data"]["total"] == 2
    resp = client.get("/enterprises/e1/risk-management/control-list", params={"zone_id": "z-other"})
    assert resp.json()["data"]["total"] == 0

    # keyword 筛选：匹配 object 或 zone
    resp = client.get("/enterprises/e1/risk-management/control-list", params={"keyword": "1#储罐"})
    assert resp.json()["data"]["total"] == 2
    resp = client.get("/enterprises/e1/risk-management/control-list", params={"keyword": "储罐区"})
    assert resp.json()["data"]["total"] == 2
    resp = client.get("/enterprises/e1/risk-management/control-list", params={"keyword": "不存在"})
    assert resp.json()["data"]["total"] == 0

    # 分页
    resp = client.get("/enterprises/e1/risk-management/control-list", params={"page": 1, "size": 1})
    data = resp.json()["data"]
    assert data["total"] == 2
    assert len(data["items"]) == 1
    resp = client.get("/enterprises/e1/risk-management/control-list", params={"page": 2, "size": 1})
    assert len(resp.json()["data"]["items"]) == 1


def test_control_list_level_filter_inherent_only(client):
    # level 筛选 inherent-only 分支：current=一般 / inherent=较大，按 level=较大 命中
    zone = RiskZone(id="z1", enterprise_id="e1", floor_id="f1", name="储罐区")
    obj = RiskObject(id="o1", enterprise_id="e1", zone_id="z1", name="1#储罐")
    obj.events = [RiskEvent(accident_type="火灾", risk_level="一般",
                            inherent_risk_level="较大")]
    zone.objects = [obj]
    db = _db(_ent(), zones=[zone])
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.get("/enterprises/e1/risk-management/control-list",
                      params={"level": "较大"})
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert data["total"] == 1
    assert data["items"][0]["accident"] == "火灾"
    assert data["items"][0]["current"] == "一般"
    assert data["items"][0]["inherent"] == "较大"


def test_control_list_other_enterprise_floor_404(client):
    # 显式传他企业 floor_id → 404「楼层不存在」
    db = _db(_ent())
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.get("/enterprises/e1/risk-management/control-list",
                      params={"floor_id": "f-other"})
    assert resp.status_code == 404
    assert resp.json()["detail"] == "楼层不存在"


# ── control-list/export ──

def test_control_list_export_xlsx(client):
    zone = _zone_with_events()
    client.app.dependency_overrides[get_db] = lambda: _db(_ent(), zones=[zone])
    resp = client.get("/enterprises/e1/risk-management/control-list/export")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert "risk_control_list.xlsx" in resp.headers["content-disposition"]
    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["风险管控清单", "等级层级汇总"]
    ws = wb.active
    assert ws.title == "风险管控清单"
    assert ws.max_row == 3  # 表头 + 2 行
    assert ws["A1"].value == "分区"


def test_control_list_export_empty(client):
    # 空清单：仅表头 1 行，双 sheet 结构完整
    db = _db(_ent(), zones=[])
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.get("/enterprises/e1/risk-management/control-list/export")
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["风险管控清单", "等级层级汇总"]
    ws = wb.active
    assert ws.title == "风险管控清单"
    assert ws.max_row == 1
    assert ws["A1"].value == "分区"


def test_control_list_export_applies_filters(client):
    # 与 control-list 同口径筛选：level 过滤后 xlsx 行数减少
    zone = _zone_with_events()
    db = _db(_ent(), zones=[zone])
    client.app.dependency_overrides[get_db] = lambda: db

    resp = client.get("/enterprises/e1/risk-management/control-list/export",
                      params={"level": "重大"})
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws.max_row == 2  # 表头 + 1 行（仅「泄漏」事件命中现有等级=重大）

    # zone_id / control_level / keyword 也透传过滤
    resp = client.get("/enterprises/e1/risk-management/control-list/export",
                      params={"zone_id": "z-other"})
    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.active.max_row == 1

    resp = client.get("/enterprises/e1/risk-management/control-list/export",
                      params={"control_level": "企业"})
    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.active.max_row == 2

    resp = client.get("/enterprises/e1/risk-management/control-list/export",
                      params={"keyword": "不存在"})
    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.active.max_row == 1


# ── risk-publicity ──

def test_risk_publicity_generates_token_and_filters_major(client):
    ent = _ent()
    zone = _zone_with_events()
    db = _db(ent, zones=[zone])
    client.app.dependency_overrides[get_db] = lambda: db

    resp = client.get("/enterprises/e1/risk-management/risk-publicity")
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert data["enterprise_name"] == "甲公司"
    assert len(data["token"]) == 64
    assert ent.public_risk_token == data["token"]
    assert len(data["items"]) == 1
    assert data["items"][0]["accident"] == "泄漏"
    assert "zone_id" not in data["items"][0]
    # 公示 items 保留 object_id（供告知卡入口链接）并含位置
    assert data["items"][0]["object_id"] == "o1"
    assert data["items"][0]["location"] == "储罐区东侧"
    # zones：四色图数据源，含双模式等级与有效色
    assert len(data["zones"]) == 1
    zone = data["zones"][0]
    assert zone["id"] == "z1"
    assert zone["floor_id"] == "f1"
    assert zone["floor_name"] is None
    assert zone["name"] == "储罐区"
    assert zone["floor_plan_polygon"] is None
    assert zone["max_level"] == "重大"
    assert zone["effective_color"] == "#ff4d4f"
    assert zone["inherent_max_level"] == "重大"
    assert zone["inherent_effective_color"] == "#ff4d4f"
    # generated_at：ISO 时间，可解析
    assert datetime.fromisoformat(data["generated_at"]).tzinfo is not None
    db.commit.assert_awaited()


def test_risk_publicity_floor_name_with_real_floor(client):
    # 分区挂真实楼层时 floor_name 非 None
    ent = _ent()
    zone = _zone_with_events()
    zone.floor = _default_floor()
    db = _db(ent, zones=[zone])
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.get("/enterprises/e1/risk-management/risk-publicity")
    assert resp.status_code == 200
    zones_data = resp.json()["data"]["zones"]
    assert len(zones_data) == 1
    assert zones_data[0]["floor_id"] == "f1"
    assert zones_data[0]["floor_name"] == "默认总图"


def test_risk_publicity_keeps_existing_token(client):
    ent = _ent(public_risk_token="existing-token")
    db = _db(ent, zones=[])
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.get("/enterprises/e1/risk-management/risk-publicity")
    assert resp.status_code == 200
    assert resp.json()["data"]["token"] == "existing-token"
    db.commit.assert_not_awaited()


def test_risk_publicity_reset_token(client):
    ent = _ent(public_risk_token="old-token")
    db = _db(ent, zones=[])
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.post("/enterprises/e1/risk-management/risk-publicity/token")
    assert resp.status_code == 200
    token = resp.json()["data"]["token"]
    assert len(token) == 64
    assert token != "old-token"
    assert ent.public_risk_token == token
    db.commit.assert_awaited()


# ── public/risk/{token} ──

def test_public_risk_404_and_desensitized(public_client):
    # 无效 token → 404「链接已失效」
    public_client.app.dependency_overrides[get_db] = lambda: _db(None)
    resp = public_client.get("/public/risk/bad")
    assert resp.status_code == 404
    assert resp.json()["detail"] == "链接已失效"


def test_public_risk_valid_token_returns_desensitized_major_items(public_client):
    ent = _ent(public_risk_token="tok-abc")
    zone = _zone_with_events()
    public_client.app.dependency_overrides[get_db] = lambda: _db(ent, zones=[zone])
    resp = public_client.get("/public/risk/tok-abc")
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert data["enterprise_name"] == "甲公司"
    assert len(data["items"]) == 1
    item = data["items"][0]
    assert item["accident"] == "泄漏"
    assert item["control_level"] == "企业"
    assert item["unit_name"] == "生产部"
    assert item["location"] == "储罐区东侧"
    assert datetime.fromisoformat(data["generated_at"]).tzinfo is not None
    # 脱敏：无 person/phone，也无内部键
    assert "person" not in item
    assert "phone" not in item
    assert "zone_id" not in item
    assert "object_id" not in item


def test_public_risk_includes_control_level_enterprise(public_client):
    ent = _ent(public_risk_token="tok-abc")
    zone = RiskZone(id="z1", enterprise_id="e1", floor_id="f1", name="储罐区")
    obj = RiskObject(id="o1", enterprise_id="e1", zone_id="z1", name="1#储罐")
    obj.events = [RiskEvent(accident_type="动火", risk_level="一般",
                            inherent_risk_level="一般", control_level="企业")]
    zone.objects = [obj]
    public_client.app.dependency_overrides[get_db] = lambda: _db(ent, zones=[zone])
    resp = public_client.get("/public/risk/tok-abc")
    assert resp.status_code == 200
    items = resp.json()["data"]["items"]
    assert len(items) == 1
    assert items[0]["accident"] == "动火"


def test_public_risk_excludes_non_major(public_client):
    ent = _ent(public_risk_token="tok-abc")
    zone = RiskZone(id="z1", enterprise_id="e1", floor_id="f1", name="储罐区")
    obj = RiskObject(id="o1", enterprise_id="e1", zone_id="z1", name="1#储罐")
    obj.events = [RiskEvent(accident_type="泄漏", risk_level="一般",
                            inherent_risk_level="一般", control_level="班组")]
    zone.objects = [obj]
    public_client.app.dependency_overrides[get_db] = lambda: _db(ent, zones=[zone])
    resp = public_client.get("/public/risk/tok-abc")
    assert resp.status_code == 200
    assert resp.json()["data"]["items"] == []
