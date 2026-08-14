import io
import json

from openpyxl import load_workbook

from app.models.enterprise_org import EnterpriseMember
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from fastapi import FastAPI, HTTPException
from fastapi.testclient import TestClient

from app.database import get_db
from app.dependencies import get_current_user
from app.models.enterprise import Enterprise
from app.models.user import User
from app.routers import enterprise_org
from app.services.enterprise_org_service import (
    IMPORT_HEADERS,
    ROLE_LABEL_MAP,
    _summarize_org_structure,
    build_member_import_template,
    normalize_org_nodes,
    parse_member_rows,
    suggest_org_tree,
    sync_org_structure,
    validate_org_tree,
)
from app.schemas.enterprise_org import OrgMember as OrgMemberSchema, OrgNode as OrgNodeSchema


def test_org_member_preserves_extra_fields_in_dump():
    member = OrgMemberSchema(name="张三", role="team_leader", phone="13800000000")
    dumped = member.model_dump()
    assert dumped["role"] == "team_leader"
    assert dumped["phone"] == "13800000000"


def test_org_node_preserves_extra_fields_in_dump():
    node = OrgNodeSchema(id="d1", type="dept", name="生产部", description="厂级部门")
    dumped = node.model_dump()
    assert dumped["description"] == "厂级部门"


def test_enterprise_member_metadata():
    assert EnterpriseMember.__tablename__ == "enterprise_members"
    cols = EnterpriseMember.__table__.columns
    assert {"id", "enterprise_id", "user_id", "org_node_id", "position", "role", "enabled"} <= set(cols.keys())


def test_enterprise_member_construct():
    m = EnterpriseMember(enterprise_id="e1", user_id="u1", role="team_leader", position="班组长")
    assert m.role == "team_leader"
    assert m.enabled is True


def test_validate_org_tree_ok():
    nodes = [
        {"id": "d1", "type": "dept", "name": "生产部", "parent_id": None, "members": []},
        {"id": "t1", "type": "team", "name": "甲班", "parent_id": "d1", "members": [{"name": "张三", "user_id": "u1"}]},
    ]
    assert validate_org_tree(nodes) == []


def test_validate_org_tree_rejects_duplicate_ids_and_bad_parent():
    nodes = [
        {"id": "d1", "type": "dept", "name": "A", "parent_id": None, "members": []},
        {"id": "d1", "type": "dept", "name": "B", "parent_id": None, "members": []},
        {"id": "t1", "type": "team", "name": "C", "parent_id": "missing", "members": []},
    ]
    errors = validate_org_tree(nodes)
    assert any("重复" in e for e in errors)
    assert any("parent" in e for e in errors)


def test_validate_org_tree_handles_non_dict_nodes():
    nodes = [None, "oops", 42, {"id": "d1", "type": "dept", "name": "A", "parent_id": None, "members": []}]
    errors = validate_org_tree(nodes)
    assert sum("必须是对象" in e for e in errors) == 3


def test_validate_org_tree_rejects_invalid_type():
    nodes = [{"id": "d1", "type": "boss", "name": "A", "parent_id": None, "members": []}]
    errors = validate_org_tree(nodes)
    assert any("type 非法" in e for e in errors)


def test_validate_org_tree_rejects_non_list_members():
    nodes = [{"id": "d1", "type": "dept", "name": "A", "parent_id": None, "members": "x"}]
    errors = validate_org_tree(nodes)
    assert any("members 必须为数组" in e for e in errors)


def test_validate_org_tree_rejects_empty_member_name():
    nodes = [{"id": "d1", "type": "dept", "name": "A", "parent_id": None, "members": [{"name": ""}]}]
    errors = validate_org_tree(nodes)
    assert any("无姓名成员" in e for e in errors)


def test_validate_org_tree_rejects_non_string_member_name():
    # 数字等非字符串 name 不再抛 TypeError，按无姓名成员拒绝
    nodes = [{"id": "d1", "type": "dept", "name": "A", "parent_id": None, "members": [{"name": 123}]}]
    errors = validate_org_tree(nodes)
    assert any("无姓名成员" in e for e in errors)


def test_validate_org_tree_handles_string_member():
    nodes = [{"id": "d1", "type": "dept", "name": "A", "parent_id": None, "members": ["张三"]}]
    errors = validate_org_tree(nodes)
    assert any("非法成员" in e for e in errors)


def test_validate_org_tree_rejects_self_parent():
    nodes = [{"id": "d1", "type": "dept", "name": "A", "parent_id": "d1", "members": []}]
    errors = validate_org_tree(nodes)
    assert any("自身" in e for e in errors)


def test_validate_org_tree_rejects_cycle():
    nodes = [
        {"id": "a", "type": "dept", "name": "A", "parent_id": "b", "members": []},
        {"id": "b", "type": "dept", "name": "B", "parent_id": "a", "members": []},
    ]
    errors = validate_org_tree(nodes)
    assert any("循环引用" in e for e in errors)


def test_validate_org_tree_accepts_normal_chain_with_shared_root():
    nodes = [
        {"id": "d1", "type": "dept", "name": "生产部", "parent_id": None, "members": []},
        {"id": "t1", "type": "team", "name": "甲班", "parent_id": "d1", "members": []},
        {"id": "t2", "type": "team", "name": "乙班", "parent_id": "d1", "members": []},
    ]
    assert validate_org_tree(nodes) == []


def test_normalize_org_nodes_generates_ids_and_defaults_members():
    nodes = [{"type": "dept", "name": "A", "parent_id": None}]
    out = normalize_org_nodes(nodes)
    assert out[0]["id"] == "node-1"
    assert out[0]["members"] == []
    assert nodes[0] == {"type": "dept", "name": "A", "parent_id": None}


def test_normalize_org_nodes_does_not_mutate_input_top_level():
    nodes = [{"id": "d1", "type": "dept", "name": "A", "parent_id": None}]
    out = normalize_org_nodes(nodes)
    out[0]["name"] = "B"
    assert nodes[0]["name"] == "A"


def test_sync_org_structure_writes_mirror():
    ent = MagicMock()
    sync_org_structure(ent, [{"id": "d1", "type": "dept", "name": "生产部", "parent_id": None, "members": []}])
    assert ent.org_structure[0]["name"] == "生产部"


# ── 组织树 + 成员 CRUD 端点测试（FastAPI + dependency_overrides + SQL 文本分发 mock，参照 test_risk_control_list.py） ──

def _org_ent(**kw):
    e = Enterprise(id="e1", user_id="u1", name="甲公司")
    for k, v in kw.items():
        setattr(e, k, v)
    return e


def _org_member(**kw):
    m = EnterpriseMember(
        id=kw.pop("id", "m1"),
        enterprise_id=kw.pop("enterprise_id", "e1"),
        user_id=kw.pop("user_id", "u2"),
        org_node_id=kw.pop("org_node_id", None),
        position=kw.pop("position", None),
        role=kw.pop("role", "member"),
        enabled=kw.pop("enabled", True),
    )
    return m


def _scalar_result(value):
    res = MagicMock()
    res.scalar_one_or_none.return_value = value
    return res


def _org_db(ent, user=None, member=None, member_rows=None):
    """按 SQL 文本特征分发：
    FROM enterprises + enterprises.user_id → 读路径归属（仅当前用户 u1 的企业）；
    FROM enterprises（无 user_id 条件）→ 写路径企业查询（路由内判 403）；
    FROM users → 用户存在性检查；
    enterprise_members + JOIN users → 成员列表 all 行；
    enterprise_members.id = → 按 id 查成员 scalar_one_or_none；
    其余 enterprise_members → 重复检查 first。
    """
    db = AsyncMock()

    def fake_add(obj):
        if isinstance(obj, EnterpriseMember) and not getattr(obj, "id", None):
            obj.id = "m1"

    db.add = MagicMock(side_effect=fake_add)

    def fake_execute(stmt, *params):
        text = str(stmt)
        if "FROM enterprises" in text:
            if "enterprises.user_id =" in text:
                owned = ent if ent and ent.user_id == "u1" else None
                return _scalar_result(owned)
            return _scalar_result(ent)
        if "FROM users" in text:
            return _scalar_result(user)
        if "FROM enterprise_members" in text:
            if "JOIN users" in text:
                res = MagicMock()
                res.all.return_value = member_rows or []
                return res
            if "enterprise_members.id =" in text:
                return _scalar_result(member)
            res = MagicMock()
            res.first.return_value = member
            return res
        return _scalar_result(None)

    db.execute.side_effect = fake_execute
    return db


@pytest.fixture()
def client():
    app = FastAPI()
    app.include_router(enterprise_org.router)

    def _override_user():
        return User(id="u1", email="a@b.c", name="A", role="admin")

    app.dependency_overrides[get_current_user] = _override_user
    app.dependency_overrides[get_db] = lambda: _org_db(_org_ent())
    with TestClient(app) as test_client:
        yield test_client


# ── GET/PUT /nodes ──

def test_org_nodes_get_returns_structure(client):
    tree = [{"id": "d1", "type": "dept", "name": "生产部", "parent_id": None,
             "members": [{"name": "张三", "user_id": "u2", "position": "部长"}]}]
    client.app.dependency_overrides[get_db] = lambda: _org_db(_org_ent(org_structure=tree))
    resp = client.get("/enterprises/e1/org/nodes")
    assert resp.status_code == 200
    assert resp.json()["data"] == tree


def test_org_nodes_get_enterprise_not_found_404(client):
    client.app.dependency_overrides[get_db] = lambda: _org_db(None)
    resp = client.get("/enterprises/e1/org/nodes")
    assert resp.status_code == 404
    assert "企业不存在" in resp.json()["detail"]


def test_org_nodes_put_valid_tree_200(client):
    ent = _org_ent(org_structure=[])
    db = _org_db(ent)
    client.app.dependency_overrides[get_db] = lambda: db
    body = {"nodes": [
        {"id": "d1", "type": "dept", "name": "生产部", "parent_id": None, "members": []},
        {"id": "t1", "type": "team", "name": "甲班", "parent_id": "d1",
         "members": [{"name": "张三", "user_id": "u2", "position": "班组长"}]},
    ]}
    resp = client.put("/enterprises/e1/org/nodes", json=body)
    assert resp.status_code == 200
    assert resp.json()["data"][0]["name"] == "生产部"
    assert ent.org_structure[1]["members"][0]["name"] == "张三"
    db.commit.assert_awaited()


def test_org_nodes_put_invalid_tree_422(client):
    ent = _org_ent()
    db = _org_db(ent)
    client.app.dependency_overrides[get_db] = lambda: db
    body = {"nodes": [
        {"id": "d1", "type": "dept", "name": "A", "parent_id": None, "members": []},
        {"id": "d1", "type": "dept", "name": "B", "parent_id": None, "members": []},
    ]}
    resp = client.put("/enterprises/e1/org/nodes", json=body)
    assert resp.status_code == 422
    detail = resp.json()["detail"]
    assert detail["code"] == "ORG_TREE_INVALID"
    assert any("重复" in e for e in detail["errors"])
    assert "重复" in detail["message"]
    db.commit.assert_not_awaited()


def test_org_nodes_put_self_loop_422(client):
    ent = _org_ent()
    db = _org_db(ent)
    client.app.dependency_overrides[get_db] = lambda: db
    body = {"nodes": [
        {"id": "d1", "type": "dept", "name": "A", "parent_id": "d1", "members": []},
    ]}
    resp = client.put("/enterprises/e1/org/nodes", json=body)
    assert resp.status_code == 422
    errors = resp.json()["detail"]["errors"]
    assert any("自身" in e for e in errors)
    db.commit.assert_not_awaited()


def test_org_nodes_put_cycle_422(client):
    ent = _org_ent()
    db = _org_db(ent)
    client.app.dependency_overrides[get_db] = lambda: db
    body = {"nodes": [
        {"id": "a", "type": "dept", "name": "A", "parent_id": "b", "members": []},
        {"id": "b", "type": "dept", "name": "B", "parent_id": "a", "members": []},
    ]}
    resp = client.put("/enterprises/e1/org/nodes", json=body)
    assert resp.status_code == 422
    errors = resp.json()["detail"]["errors"]
    assert any("循环引用" in e for e in errors)
    db.commit.assert_not_awaited()


def test_org_nodes_put_preserves_extra_fields(client):
    ent = _org_ent(org_structure=[])
    db = _org_db(ent)
    client.app.dependency_overrides[get_db] = lambda: db
    body = {"nodes": [
        {"id": "d1", "type": "dept", "name": "生产部", "parent_id": None,
         "description": "厂级部门",
         "members": [{"name": "张三", "role": "team_leader", "phone": "13800000000"}]},
    ]}
    resp = client.put("/enterprises/e1/org/nodes", json=body)
    assert resp.status_code == 200
    saved = ent.org_structure[0]
    assert saved["description"] == "厂级部门"
    assert saved["members"][0]["role"] == "team_leader"
    assert saved["members"][0]["phone"] == "13800000000"


def test_org_nodes_put_non_owner_403(client):
    client.app.dependency_overrides[get_db] = lambda: _org_db(_org_ent(user_id="u2"))
    body = {"nodes": [{"id": "d1", "type": "dept", "name": "A", "parent_id": None, "members": []}]}
    resp = client.put("/enterprises/e1/org/nodes", json=body)
    assert resp.status_code == 403


# ── POST /members ──

def test_members_post_created_with_default_role(client):
    ent = _org_ent()
    user = User(id="u2", email="b@b.c", name="张三", role="user")
    db = _org_db(ent, user=user)
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.post("/enterprises/e1/org/members", json={"user_id": "u2"})
    assert resp.status_code == 201
    data = resp.json()["data"]
    assert data["role"] == "member"
    assert data["user_id"] == "u2"
    assert data["enterprise_id"] == "e1"
    added = db.add.call_args[0][0]
    assert isinstance(added, EnterpriseMember)
    assert added.role == "member"
    assert added.enabled is True
    db.commit.assert_awaited()


def test_members_post_user_not_found_404(client):
    ent = _org_ent()
    db = _org_db(ent, user=None)
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.post("/enterprises/e1/org/members", json={"user_id": "nobody"})
    assert resp.status_code == 404
    assert "用户不存在" in resp.json()["detail"]


def test_members_post_duplicate_409(client):
    ent = _org_ent()
    user = User(id="u2", email="b@b.c", name="张三", role="user")
    db = _org_db(ent, user=user, member=_org_member())
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.post("/enterprises/e1/org/members", json={"user_id": "u2"})
    assert resp.status_code == 409
    assert "已是企业成员" in resp.json()["detail"]


def test_members_post_non_owner_403(client):
    ent = _org_ent(user_id="u2")
    user = User(id="u3", email="c@b.c", name="李四", role="user")
    client.app.dependency_overrides[get_db] = lambda: _org_db(ent, user=user)
    resp = client.post("/enterprises/e1/org/members", json={"user_id": "u3"})
    assert resp.status_code == 403


# ── PUT /members ──

def test_members_put_updates_fields_without_clearing_others(client):
    ent = _org_ent()
    member = _org_member(org_node_id="n1", position="旧岗位", role="member", enabled=True)
    db = _org_db(ent, member=member)
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.put("/enterprises/e1/org/members/m1", json={"position": "新岗位"})
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert data["position"] == "新岗位"
    assert data["org_node_id"] == "n1"  # exclude_unset：未传字段不误清
    assert member.role == "member"
    assert member.enabled is True
    db.commit.assert_awaited()


def test_members_put_explicit_null_role_rejected_422(client):
    ent = _org_ent()
    member = _org_member(role="member", enabled=True)
    db = _org_db(ent, member=member)
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.put("/enterprises/e1/org/members/m1", json={"role": None})
    assert resp.status_code == 422
    assert "role 不能为 null" in resp.json()["detail"]
    assert member.role == "member"  # 未落库
    db.commit.assert_not_awaited()


def test_members_put_explicit_null_enabled_rejected_422(client):
    ent = _org_ent()
    member = _org_member(role="member", enabled=True)
    db = _org_db(ent, member=member)
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.put("/enterprises/e1/org/members/m1", json={"enabled": None})
    assert resp.status_code == 422
    assert "enabled 不能为 null" in resp.json()["detail"]
    assert member.enabled is True
    db.commit.assert_not_awaited()


def test_members_put_explicit_null_position_clears(client):
    ent = _org_ent()
    member = _org_member(org_node_id="n1", position="旧岗位", role="member", enabled=True)
    db = _org_db(ent, member=member)
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.put("/enterprises/e1/org/members/m1", json={"position": None})
    assert resp.status_code == 200
    assert resp.json()["data"]["position"] is None
    assert member.position is None  # 显式 null 保留清空语义
    assert member.role == "member"
    assert member.enabled is True
    db.commit.assert_awaited()


def test_members_put_404(client):
    ent = _org_ent()
    db = _org_db(ent, member=None)
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.put("/enterprises/e1/org/members/missing", json={"position": "新岗位"})
    assert resp.status_code == 404
    assert "成员不存在" in resp.json()["detail"]


def test_members_put_non_owner_403(client):
    client.app.dependency_overrides[get_db] = lambda: _org_db(_org_ent(user_id="u2"))
    resp = client.put("/enterprises/e1/org/members/m1", json={"position": "新岗位"})
    assert resp.status_code == 403


# ── DELETE /members ──

def test_members_delete_success(client):
    ent = _org_ent()
    member = _org_member()
    db = _org_db(ent, member=member)
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.delete("/enterprises/e1/org/members/m1")
    assert resp.status_code == 200
    body = resp.json()
    assert body["code"] == 0
    assert body["data"] is None
    assert body["message"] == "已删除"
    db.delete.assert_awaited_with(member)
    db.commit.assert_awaited()


def test_members_delete_404(client):
    ent = _org_ent()
    db = _org_db(ent, member=None)
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.delete("/enterprises/e1/org/members/missing")
    assert resp.status_code == 404
    assert "成员不存在" in resp.json()["detail"]


def test_members_delete_non_owner_403(client):
    client.app.dependency_overrides[get_db] = lambda: _org_db(_org_ent(user_id="u2"))
    resp = client.delete("/enterprises/e1/org/members/m1")
    assert resp.status_code == 403


# ── GET /members ──

def test_members_get_joins_email_and_name(client):
    ent = _org_ent()
    member = _org_member(user_id="u2", org_node_id="n1", position="班组长", role="team_leader")
    user = User(id="u2", email="b@b.c", name="张三", role="user")
    db = _org_db(ent, member_rows=[(member, user)])
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.get("/enterprises/e1/org/members")
    assert resp.status_code == 200
    items = resp.json()["data"]
    assert len(items) == 1
    item = items[0]
    assert item["id"] == "m1"
    assert item["enterprise_id"] == "e1"
    assert item["user_id"] == "u2"
    assert item["email"] == "b@b.c"
    assert item["name"] == "张三"
    assert item["org_node_id"] == "n1"
    assert item["position"] == "班组长"
    assert item["role"] == "team_leader"
    assert item["enabled"] is True


# ── GET /members/search（按邮箱搜索可绑定账号）与 GET /members/template（模板下载） ──

def test_members_search_returns_bindable_users(client):
    ent = _org_ent()
    user = User(id="u2", email="zhang@x.com", name="张三", role="user")
    db = _org_db(ent, user=user, member=None)
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.get("/enterprises/e1/org/members/search", params={"email": "zhang"})
    assert resp.status_code == 200
    assert resp.json()["data"] == [{"id": "u2", "email": "zhang@x.com", "name": "张三"}]


def test_members_search_excludes_existing_members(client):
    ent = _org_ent()
    user = User(id="u2", email="zhang@x.com", name="张三", role="user")
    db = _org_db(ent, user=user, member=_org_member(user_id="u2"))
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.get("/enterprises/e1/org/members/search", params={"email": "zhang"})
    assert resp.status_code == 200
    assert resp.json()["data"] == []


def test_members_search_empty_email_returns_empty(client):
    ent = _org_ent()
    db = _org_db(ent)
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.get("/enterprises/e1/org/members/search", params={"email": "  "})
    assert resp.status_code == 200
    assert resp.json()["data"] == []


def test_members_search_non_owner_404(client):
    ent = _org_ent(user_id="u2")
    client.app.dependency_overrides[get_db] = lambda: _org_db(ent)
    resp = client.get("/enterprises/e1/org/members/search", params={"email": "zhang"})
    assert resp.status_code == 404
    assert "企业不存在" in resp.json()["detail"]


def test_members_template_downloads_xlsx(client):
    ent = _org_ent()
    client.app.dependency_overrides[get_db] = lambda: _org_db(ent)
    resp = client.get("/enterprises/e1/org/members/template")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert "member_import_template.xlsx" in resp.headers["content-disposition"]
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert [ws.cell(row=1, column=c).value for c in range(1, 7)] == IMPORT_HEADERS


def test_members_template_non_owner_404(client):
    ent = _org_ent(user_id="u2")
    client.app.dependency_overrides[get_db] = lambda: _org_db(ent)
    resp = client.get("/enterprises/e1/org/members/template")
    assert resp.status_code == 404
    assert "企业不存在" in resp.json()["detail"]


# ── Excel 导入 + 责任人选择器（任务 4） ──

def _import_xlsx_bytes(rows):
    """用模板生成 xlsx 字节：rows 为表头后的数据行列表。"""
    wb = build_member_import_template()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ── 服务层：模板 + 解析 ──

def test_build_member_import_template_has_headers_and_role_dropdown():
    wb = build_member_import_template()
    ws = wb.active
    assert [ws.cell(row=1, column=c).value for c in range(1, 7)] == ["姓名", "邮箱", "部门", "班组", "岗位", "角色"]
    dvs = ws.data_validations.dataValidation
    assert dvs
    assert any("企业管理员" in (dv.formula1 or "") and "班组长" in (dv.formula1 or "") and "员工" in (dv.formula1 or "")
               for dv in dvs)


def test_parse_member_rows_ok():
    rows = [{"姓名": "张三", "邮箱": "zhang@x.com", "部门": "生产部", "班组": "甲班", "岗位": "班组长", "角色": "班组长"}]
    parsed = parse_member_rows(rows)
    assert parsed[0]["name"] == "张三"
    assert parsed[0]["email"] == "zhang@x.com"
    assert parsed[0]["department"] == "生产部"
    assert parsed[0]["team"] == "甲班"
    assert parsed[0]["position"] == "班组长"
    assert parsed[0]["role"] == "team_leader"
    assert "error" not in parsed[0]


def test_parse_member_rows_role_mapping():
    assert ROLE_LABEL_MAP == {"企业管理员": "enterprise_admin", "班组长": "team_leader", "员工": "member"}
    rows = [
        {"姓名": "张三", "邮箱": "zhang@x.com", "角色": "企业管理员"},
        {"姓名": "李四", "邮箱": "li@x.com", "角色": "班组长"},
        {"姓名": "王五", "邮箱": "wang@x.com", "角色": "员工"},
        {"姓名": "赵六", "邮箱": "zhao@x.com", "角色": "未知角色"},
        {"姓名": "孙七", "邮箱": "sun@x.com", "角色": ""},
    ]
    parsed = parse_member_rows(rows)
    assert [p["role"] for p in parsed] == ["enterprise_admin", "team_leader", "member", "member", "member"]


def test_parse_member_rows_missing_email_error():
    parsed = parse_member_rows([{"姓名": "张三", "邮箱": "  ", "角色": "班组长"}])
    assert parsed[0]["email"] == ""
    assert parsed[0]["error"] == "邮箱必填"
    assert parsed[0]["role"] == "team_leader"


def test_parse_member_rows_invalid_email_error():
    parsed = parse_member_rows([{"姓名": "张三", "邮箱": "zhangx.com", "角色": "员工"}])
    assert "邮箱格式" in parsed[0]["error"]
    assert parsed[0]["role"] == "member"


# ── POST /members/import ──

def test_members_import_success_creates_nodes_and_member(client):
    ent = _org_ent(org_structure=[])
    user = User(id="u2", email="zhang@x.com", name="张三", role="user")
    db = _org_db(ent, user=user, member=None)
    client.app.dependency_overrides[get_db] = lambda: db
    content = _import_xlsx_bytes([
        ["张三", "zhang@x.com", "生产部", "甲班", "班组长", "班组长"],
    ])
    resp = client.post(
        "/enterprises/e1/org/members/import",
        files={"file": ("members.xlsx", content, _XLSX_MIME)},
    )
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert data["imported"] == 1
    assert data["skipped"] == 0
    assert data["errors"] == []
    # org_structure 按部门/班组名创建节点，id 复用 normalize 短 id 规则
    dept = ent.org_structure[0]
    assert dept["name"] == "生产部"
    assert dept["type"] == "dept"
    assert dept["id"].startswith("node-")
    team = next(n for n in ent.org_structure if n["type"] == "team")
    assert team["name"] == "甲班"
    assert team["parent_id"] == dept["id"]
    added = [c.args[0] for c in db.add.call_args_list if isinstance(c.args[0], EnterpriseMember)]
    assert len(added) == 1
    assert added[0].user_id == "u2"
    assert added[0].role == "team_leader"
    assert added[0].org_node_id == team["id"]
    db.commit.assert_awaited()


def test_members_import_user_not_found_error_row(client):
    ent = _org_ent(org_structure=[])
    db = _org_db(ent, user=None, member=None)
    client.app.dependency_overrides[get_db] = lambda: db
    content = _import_xlsx_bytes([
        ["张三", "nobody@x.com", "生产部", "甲班", "班组长", "班组长"],
    ])
    resp = client.post(
        "/enterprises/e1/org/members/import",
        files={"file": ("members.xlsx", content, _XLSX_MIME)},
    )
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert data["imported"] == 0
    assert len(data["errors"]) == 1
    assert data["errors"][0]["row"] == 2
    assert "用户不存在" in data["errors"][0]["reason"]
    db.commit.assert_awaited()


def test_members_import_duplicate_skipped(client):
    ent = _org_ent(org_structure=[])
    user = User(id="u2", email="zhang@x.com", name="张三", role="user")
    db = _org_db(ent, user=user, member=_org_member(user_id="u2"))
    client.app.dependency_overrides[get_db] = lambda: db
    content = _import_xlsx_bytes([
        ["张三", "zhang@x.com", "生产部", "甲班", "班组长", "班组长"],
    ])
    resp = client.post(
        "/enterprises/e1/org/members/import",
        files={"file": ("members.xlsx", content, _XLSX_MIME)},
    )
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert data["imported"] == 0
    assert data["skipped"] == 1
    assert data["errors"] == []
    db.commit.assert_awaited()


def test_members_import_duplicate_email_within_file_skipped(client):
    # 文件内重复邮箱：请求内去重（提交前 DB 查不到本批未 flush 成员）
    ent = _org_ent(org_structure=[])
    user = User(id="u2", email="zhang@x.com", name="张三", role="user")
    db = _org_db(ent, user=user, member=None)
    client.app.dependency_overrides[get_db] = lambda: db
    content = _import_xlsx_bytes([
        ["张三", "zhang@x.com", "生产部", "甲班", "班组长", "班组长"],
        ["张三", "zhang@x.com", "生产部", "甲班", "班组长", "班组长"],
    ])
    resp = client.post(
        "/enterprises/e1/org/members/import",
        files={"file": ("members.xlsx", content, _XLSX_MIME)},
    )
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert data["imported"] == 1
    assert data["skipped"] == 1
    assert data["errors"] == []
    db.commit.assert_awaited()


def test_members_import_invalid_email_error_row(client):
    ent = _org_ent(org_structure=[])
    db = _org_db(ent, user=None, member=None)
    client.app.dependency_overrides[get_db] = lambda: db
    content = _import_xlsx_bytes([
        ["张三", "bad-email", "生产部", "甲班", "班组长", "班组长"],
    ])
    resp = client.post(
        "/enterprises/e1/org/members/import",
        files={"file": ("members.xlsx", content, _XLSX_MIME)},
    )
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert data["imported"] == 0
    assert len(data["errors"]) == 1
    assert data["errors"][0]["row"] == 2
    assert "邮箱" in data["errors"][0]["reason"]
    db.commit.assert_awaited()


def test_members_import_non_owner_403(client):
    ent = _org_ent(user_id="u2", org_structure=[])
    client.app.dependency_overrides[get_db] = lambda: _org_db(ent)
    content = _import_xlsx_bytes([
        ["张三", "zhang@x.com", "生产部", "甲班", "班组长", "班组长"],
    ])
    resp = client.post(
        "/enterprises/e1/org/members/import",
        files={"file": ("members.xlsx", content, _XLSX_MIME)},
    )
    assert resp.status_code == 403


def test_members_import_corrupt_file_400(client):
    ent = _org_ent(org_structure=[])
    client.app.dependency_overrides[get_db] = lambda: _org_db(ent)
    resp = client.post(
        "/enterprises/e1/org/members/import",
        files={"file": ("members.xlsx", b"not an xlsx file", _XLSX_MIME)},
    )
    assert resp.status_code == 400
    assert "导入文件格式无效" in resp.json()["detail"]


def test_members_import_header_mismatch_400(client):
    from openpyxl import Workbook

    ent = _org_ent(org_structure=[])
    client.app.dependency_overrides[get_db] = lambda: _org_db(ent)
    wb = Workbook()
    ws = wb.active
    ws.append(["姓名", "邮箱", "部门", "班组", "岗位", "备注"])
    ws.append(["张三", "zhang@x.com", "生产部", "甲班", "班组长", "x"])
    buf = io.BytesIO()
    wb.save(buf)
    resp = client.post(
        "/enterprises/e1/org/members/import",
        files={"file": ("members.xlsx", buf.getvalue(), _XLSX_MIME)},
    )
    assert resp.status_code == 400
    assert "表头与模板不符" in resp.json()["detail"]


def test_members_import_header_order_and_whitespace_insensitive(client):
    from openpyxl import Workbook

    ent = _org_ent(org_structure=[])
    user = User(id="u2", email="zhang@x.com", name="张三", role="user")
    db = _org_db(ent, user=user, member=None)
    client.app.dependency_overrides[get_db] = lambda: db
    wb = Workbook()
    ws = wb.active
    ws.append([" 邮箱 ", "姓名", "角色", "岗位", "班组", "部门"])
    ws.append(["zhang@x.com", "张三", "班组长", "班组长", "甲班", "生产部"])
    buf = io.BytesIO()
    wb.save(buf)
    resp = client.post(
        "/enterprises/e1/org/members/import",
        files={"file": ("members.xlsx", buf.getvalue(), _XLSX_MIME)},
    )
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert data["imported"] == 1
    assert data["errors"] == []
    assert ent.org_structure[0]["name"] == "生产部"


def test_members_import_file_too_large_413(client):
    ent = _org_ent(org_structure=[])
    client.app.dependency_overrides[get_db] = lambda: _org_db(ent)
    oversized = b"x" * (5 * 1024 * 1024 + 1)
    resp = client.post(
        "/enterprises/e1/org/members/import",
        files={"file": ("members.xlsx", oversized, _XLSX_MIME)},
    )
    assert resp.status_code == 413
    assert "5MB" in resp.json()["detail"]


# ── GET /members/available ──

def test_members_available_returns_enabled_with_org_path(client):
    tree = [
        {"id": "d1", "type": "dept", "name": "生产部", "parent_id": None, "members": []},
        {"id": "t1", "type": "team", "name": "甲班", "parent_id": "d1", "members": []},
    ]
    ent = _org_ent(org_structure=tree)
    member = _org_member(user_id="u2", org_node_id="t1", position="班组长", role="team_leader", enabled=True)
    user = User(id="u2", email="zhang@x.com", name="张三", role="user")
    db = _org_db(ent, member_rows=[(member, user)])
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.get("/enterprises/e1/org/members/available")
    assert resp.status_code == 200
    items = resp.json()["data"]
    assert len(items) == 1
    item = items[0]
    assert item["id"] == "m1"
    assert item["name"] == "张三"
    assert item["email"] == "zhang@x.com"
    assert item["role"] == "team_leader"
    assert item["position"] == "班组长"
    assert item["org_path"] == "生产部/甲班"


def test_members_available_filters_enabled_and_missing_node_path(client):
    ent = _org_ent(org_structure=[])
    member = _org_member(user_id="u2", org_node_id=None, position="安全员", role="member", enabled=True)
    user = User(id="u2", email="zhang@x.com", name="张三", role="user")
    db = _org_db(ent, member_rows=[(member, user)])
    client.app.dependency_overrides[get_db] = lambda: db
    resp = client.get("/enterprises/e1/org/members/available")
    assert resp.status_code == 200
    items = resp.json()["data"]
    assert len(items) == 1
    assert items[0]["org_path"] == ""
    # SQL 层带 enabled 过滤（停用成员不被选中）
    stmts = [str(c.args[0]) for c in db.execute.call_args_list]
    members_stmt = next(s for s in stmts if "FROM enterprise_members" in s and "JOIN users" in s)
    assert "enabled" in members_stmt


def test_members_available_non_owner_read_404(client):
    # 读路径沿用现有语义：当前无成员归属关系，非企业主视为企业不存在 → 404
    ent = _org_ent(user_id="u2", org_structure=[])
    client.app.dependency_overrides[get_db] = lambda: _org_db(ent)
    resp = client.get("/enterprises/e1/org/members/available")
    assert resp.status_code == 404
    assert "企业不存在" in resp.json()["detail"]


# ── AI 建树建议（任务 5，文本通道，mock LLM） ──

def test_summarize_org_structure_paths():
    nodes = [
        {"id": "d1", "type": "dept", "name": "生产部", "parent_id": None, "members": []},
        {"id": "t1", "type": "team", "name": "甲班", "parent_id": "d1", "members": []},
        {"id": "t2", "type": "team", "name": "乙班", "parent_id": "d1", "members": []},
    ]
    summary = _summarize_org_structure(nodes)
    assert "生产部" in summary
    assert "生产部/甲班" in summary
    assert "生产部/乙班" in summary


def test_summarize_org_structure_cycle_safe():
    nodes = [
        {"id": "a", "type": "dept", "name": "A", "parent_id": "b", "members": []},
        {"id": "b", "type": "dept", "name": "B", "parent_id": "a", "members": []},
    ]
    summary = _summarize_org_structure(nodes)
    assert isinstance(summary, str)
    assert summary != ""


def test_summarize_org_structure_empty():
    assert _summarize_org_structure([]) == "（暂无）"


@pytest.mark.asyncio
async def test_ai_suggest_org_tree_ok():
    fake = {"nodes": [
        {"id": "d1", "type": "dept", "name": "生产部", "parent_id": None, "members": [{"name": "张三", "position": "班组长"}]},
        {"id": "t1", "type": "team", "name": "甲班", "parent_id": "d1", "members": []},
    ]}
    with patch("app.services.enterprise_org_service.llm_text_completion",
               AsyncMock(return_value=json.dumps(fake, ensure_ascii=False))):
        out = await suggest_org_tree({"industry": "化工", "employee_count": 120}, None)
    assert out["available"] is True
    assert out["nodes"][0]["type"] == "dept"


@pytest.mark.asyncio
async def test_ai_suggest_org_tree_fallback():
    with patch("app.services.enterprise_org_service.llm_text_completion",
               AsyncMock(side_effect=Exception("timeout"))):
        out = await suggest_org_tree({"industry": "化工"}, None)
    assert out["available"] is False


# ── POST /org/ai-suggest 端点 ──

def test_ai_suggest_endpoint_returns_result(client):
    ent = _org_ent(industry="化工", employee_count=120, org_structure=[])
    db = _org_db(ent)
    client.app.dependency_overrides[get_db] = lambda: db
    fake_result = {"available": True, "nodes": [
        {"id": "d1", "type": "dept", "name": "生产部", "parent_id": None, "members": []},
    ]}
    with patch("app.routers.enterprise_org._get_ai_config", AsyncMock(return_value=MagicMock())), \
         patch("app.routers.enterprise_org.suggest_org_tree", AsyncMock(return_value=fake_result)) as mock_suggest:
        resp = client.post("/enterprises/e1/org/ai-suggest")
    assert resp.status_code == 200
    assert resp.json()["data"] == fake_result
    assert mock_suggest.await_count == 1
    info = mock_suggest.await_args.args[0]
    assert info["industry"] == "化工"
    assert info["employee_count"] == 120


def test_ai_suggest_endpoint_no_ai_config_still_200(client):
    ent = _org_ent(industry="化工")
    db = _org_db(ent)
    client.app.dependency_overrides[get_db] = lambda: db
    fallback = {"available": False, "note": "AI 不可用，请手动维护组织架构"}
    with patch("app.routers.enterprise_org._get_ai_config",
               AsyncMock(side_effect=HTTPException(400, "系统未配置 AI 模型，请联系管理员"))), \
         patch("app.routers.enterprise_org.suggest_org_tree", AsyncMock(return_value=fallback)) as mock_suggest:
        resp = client.post("/enterprises/e1/org/ai-suggest")
    assert resp.status_code == 200
    assert resp.json()["data"]["available"] is False
    # 未配置 AI 时配置转 None 传给服务兜底
    assert mock_suggest.await_args.args[1] is None


def test_ai_suggest_endpoint_non_owner_403(client):
    ent = _org_ent(user_id="u2")
    client.app.dependency_overrides[get_db] = lambda: _org_db(ent)
    resp = client.post("/enterprises/e1/org/ai-suggest")
    assert resp.status_code == 403
